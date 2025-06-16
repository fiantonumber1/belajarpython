from tqi_module import generate_pairs,export_xlsx_to_pdf,center_columns, get_valid_block_range,wrapsheet,titlecreate,tqi_kelas,font_set,transform_name,upordown,databasekota,segment_length_function
import re 
import pandas as pd
import numpy as np
import PySimpleGUI as sg
import chardet
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image    
from openpyxl.styles import Alignment
from pathlib import Path

# Sample per meter didefinisikan setiap 4 baris didefenisikan sebagai 1 meter
SAMPLES_PER_METER = 4  # Jumlah sampel per meter

# Meter capture values adalah nilai yang ingin dinilai misal nanti ingin check tqi dari suatu range misal tiap 40 meter 1 nilai tqi pilih opsi 40
meter_capture_values = [str(i) for i in range(20, 101, 5)]  # List: "20", "25", "30", ..., "100"
meter_capture_values.extend([str(200), str(1000)])

# Kelas jalur yang digunakan untuk klasifikasi
track_classes = ['I', 'II', 'III', 'IV']


isZeroValid=False  # Variabel untuk menentukan apakah NaN dianggap valid atau tidak


# Fungsi untuk memproses file CSV dan menyimpan hasilnya ke file Excel
def process_csv(file_path, selected_headers, block_size, origin_kilometer, encoding,idinput):

    startpoint = ""
    endpoint = ""
    kotadilewati = []
    pairs = generate_pairs(selected_headers)

    try:
        trend = "up"
        
        start_row, rows_to_read = get_valid_block_range(file_path, selected_headers, encoding)
        rows_to_skip = [0] + list(range(2, start_row + 2))  # Mulai dari 2 sampai start_row + 1
        data = pd.read_csv(file_path, sep=',', encoding=encoding, skiprows=rows_to_skip, nrows=rows_to_read)
        
        # Cek apakah trend naik atau turun
        before = data.iloc[0]['Km']
        after = data.iloc[1]['Km']
        trend = upordown(before, after)

        lines = data.iloc[2]['Line']

        def loopcheck(data, startloop):
            # """
            #Mengambil informasi kilometer dari baris data tertentu.
            #
            #Parameters:
            #    data (DataFrame): Data berisi kolom 'Km'.
            #    startloop (int): Indeks baris saat ini.
            #
            #Returns:
            #    tuple: (origin_kilometer, meter_start, checkvalue, startloop)
            #

            # Ambil nilai kilometer bulat
            startloop += 1
            if startloop >= len(data):
                raise IndexError("Index `start` melebihi jumlah baris pada data.")
            origin_kilometer =  int(data.iloc[startloop]['Km']) # kilometer
            meter_start = int(data.iloc[startloop]['Km']*1000 - int(data.iloc[startloop]['Km']) * 1000)
            checkvalue = int(data.iloc[startloop]['Km'] * 10000) % 10
            return origin_kilometer, meter_start, checkvalue, startloop

        meter_start, checkvalue, startloop = 0, 0, -1

        if origin_kilometer == "":
            try:
                origin_kilometer, meter_start, checkvalue, startloop = loopcheck(data, startloop)
                max_iterations = len(data)
                while not (1.9 <= checkvalue <= 3.1) and startloop < max_iterations:
                    origin_kilometer, meter_start, checkvalue, startloop = loopcheck(data, startloop)
                    print("Ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
                if startloop >= max_iterations:
                    print("Tidak ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
            except IndexError:
                print("Tidak ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
        else:
            origin_kilometer = float(origin_kilometer)

        rows_to_skip = [0] + list(range(2, start_row + 2+startloop))  # Mulai dari 2 sampai start_row + 1

        data = pd.read_csv(file_path, sep=',', encoding=encoding, skiprows=rows_to_skip,nrows=rows_to_read-startloop)

        error_string = 'raise error'
        error_rows = data[data.apply(lambda row: row.astype(str).str.contains(error_string).any(), axis=1)].index.tolist()
        if error_rows:
            raise ValueError(f"Data mengandung string '{error_string}' pada baris ke: {error_rows}")

        headers = data.columns.tolist()
        print(headers)

        # Filter kolom berdasarkan header yang dipilih
        selected_columns = data[selected_headers]

        # Validasi nilai block_size
        if block_size <= 0:
            raise ValueError("block_size harus berupa bilangan positif.")

        # Validasi data kosong setelah filter
        if selected_columns.empty:
            raise ValueError("Data kosong setelah filtering berdasarkan header yang dipilih.")

        # Proses data sesuai kebutuhan

        # Menghapus baris pertama dari data untuk isi sheet
        data_to_save = selected_columns.iloc[1:]

        # Menyimpan hasil ke file Excel baru
        output_file = "allparameter.xlsx"
        stdev_summary = {}  # Untuk menyimpan nilai standar deviasi setiap sheet

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Menyimpan semua kolom ke satu sheet bernama "AllParameter"
            data_to_save.to_excel(writer, index=False, header=True, sheet_name="AllParameter")

            # Definisi parameter dan pasangan berdasarkan nama
            parameters = []
            # Menyimpan setiap kolom ke sheet terpisah dengan nama sesuai header
            for idx, header in enumerate(selected_headers):
               
                sheet_name_cleaned = re.sub(r'\[.*?\]', '', header)  # Menghapus teks dalam tanda kurung siku
                sheet_name_cleaned = re.sub(r'[.\-]', '', sheet_name_cleaned).strip()  # Menghapus titik dan tanda minus

                # Potong nama sheet jika lebih dari 31 karakter
                sheet_name = sheet_name_cleaned[:31] if sheet_name_cleaned else f"Sheet{idx + 1}"

                sheet_name = transform_name(sheet_name)

                parameters.append(sheet_name)

                selected_column_data = data.iloc[1:, [data.columns.get_loc(header)]]  # Data dari baris ke-2 dan seterusnya

                # Membagi data menjadi blok dengan block_size baris
                total_rows = len(selected_column_data)
                blocks = [selected_column_data.iloc[i:i + block_size] for i in range(0, total_rows, block_size)]

                # Membuat DataFrame untuk setiap blok
                max_length = max(len(block) for block in blocks)
                split_data = pd.DataFrame()

                for i, block in enumerate(blocks):
                    # Setiap blok menjadi kolom baru
                    column_data = block.values.flatten()
                    if len(column_data) < max_length:
                        # Isi kekurangan baris dengan NaN
                        column_data = list(column_data) + [None] * (max_length - len(column_data))
                    split_data[f"Part{i + 1}"] = column_data

                # Menghitung standar deviasi untuk setiap kolom
                if(isZeroValid):
                    # Jika isZeroValid adalah true maka nol diakui, jika nol diakui maka tidak ada NaN maka semua dihitung
                    stdev_row = split_data.apply(lambda x: np.nanstd(x), axis=0)  # Menggunakan np.nanstd untuk menghitung standar deviasi
                else:
                    # Jika isZeroValid adalah false maka nol tidak diakui, jika nol tidak diakui maka ada NaN maka Nan didrop
                    stdev_row = split_data.apply(lambda x: np.std(x.dropna()), axis=0)


                # Menyimpan standar deviasi dalam dictionary
                stdev_summary[sheet_name] = stdev_row.values

                # Menambahkan baris standar deviasi ke DataFrame
                split_data.loc[max_length] = stdev_row

                # Menyimpan data yang telah dipecah ke sheet Excel
                split_data.to_excel(writer, index=False, header=True, sheet_name=sheet_name)

            summary_df = pd.DataFrame()

            # Menambahkan data dari stdev_summary ke summary_df
            for sheet_name, stdev_values in stdev_summary.items():
                # Pastikan stdev_values adalah array
                if isinstance(stdev_values, (np.ndarray, list)):
                    summary_df[sheet_name] = stdev_values

            # Menyimpan summary ke sheet "Summary"
            summary_df.to_excel(writer, index=False, header=True, sheet_name="Summary")


            # Inisialisasi kolom total
            total_column = []
            total_column_2 = []

            # Membersihkan setiap elemen dalam pairs
            pairs_cleaned = [
                (
                    transform_name(re.sub(r'\[.*?\]', '', param1)),
                    transform_name(re.sub(r'\[.*?\]', '', param2))   
                 )
                for param1, param2 in pairs
            ]

            pairs = [
                (re.sub(r'[.\-]', '', param1).strip(), re.sub(r'[.\-]', '', param2).strip())
                for param1, param2 in pairs_cleaned
            ]

            all_params = []
            for pair in pairs:
                for param in pair:
                    all_params.append(param)

            # Menghilangkan duplikasi
            paired_parameters = list(set(all_params))


            

            # Memastikan remaining_params benar
            remaining_params = [param for param in parameters if param not in paired_parameters]

            # Dictionary untuk indeks parameter
            param_indices = {param: idx for idx, param in enumerate(parameters)}

            
            
            # Iterasi untuk menghitung total
            for idx in range(len(summary_df)):
                total_kai = 0
                total_en = 0

                # # Menghitung total untuk pasangan parameter
                for param1_name, param2_name in pairs:
                    if param1_name in param_indices and param2_name in param_indices:
                        param1_idx = param_indices[param1_name]
                        param2_idx = param_indices[param2_name]
                        avg_pair = (summary_df.iloc[idx, param1_idx] + summary_df.iloc[idx, param2_idx]) / 2

                        total_kai += avg_pair
                        total_en += avg_pair**2

                # Menghitung total untuk sisa parameter
                for param_name in remaining_params:
                    param_idx = param_indices[param_name]
                    total_kai += summary_df.iloc[idx, param_idx]
                    total_en += summary_df.iloc[idx, param_idx]**2

                # Menyimpan hasil
                sqrtotal_en = round(total_en**0.5, 2)



                total_column_2.append(sqrtotal_en)
                total_column.append(round(total_kai, 2))
            
             # Menambahkan kolom total ke DataFrame summary_df
            summary_df['TQI Total KAI'] = total_column
            summary_df['TQI Total EN'] = total_column_2


            # Menambahkan kolom total ke DataFrame dengan aturan khusus
            summary_df['Track Number'] = [lines] * len(summary_df)

            # segment length adalah panjang segmen dalam 1 block data stdev dan memiliki satuan meter
            # Menghitung panjang segmen berdasarkan block_size, misal SAMPLES_PER_METER adalah 4 dan block_size adalah 160 maka segment_length adalah 40 meter berarti 1 block data stdev adalah 40 meter
            segment_length = segment_length_function(block_size, SAMPLES_PER_METER)
            

            # Menambahkan kolom "Meter Section" dengan aturan khusus
            origin_kilometer_to_meter = origin_kilometer * 1000  # Mengonversi kilometer ke meter
            if(trend=="up"):
                summary_df['Meter Awal'] = origin_kilometer_to_meter + meter_start + summary_df.index * segment_length   # satuan meter
            elif(trend=="down"):
                summary_df['Meter Awal'] = origin_kilometer_to_meter + meter_start - summary_df.index * segment_length  # satuan meter

            # Menambahkan kolom "Track Length" dengan aturan khusus
            summary_df['Track Length'] = [segment_length] * len(summary_df)


             # Menggabungkan kolom 'LPROF_MID' dan 'RPROF_MID' dengan menghitung rata-ratanya
            if 'Meter Awal' in summary_df.columns and 'Track Length' in summary_df.columns:
                if(trend=="up"):
                    summary_df['Meter Akhir'] = (summary_df['Meter Awal'] + summary_df['Track Length']) 
                elif(trend=="down"):
                    summary_df['Meter Akhir'] = (summary_df['Meter Awal'] - summary_df['Track Length']) 

            if 'Meter Awal' in summary_df.columns:
                summary_df['Km Awal'] = np.floor(summary_df['Meter Awal'] / 1000).astype(int) 

            if 'Meter Akhir' in summary_df.columns:
                summary_df['Km Akhir'] = np.floor(summary_df['Meter Akhir'] / 1000).astype(int) 


            if 'Meter Awal' in summary_df.columns:
                summary_df['Meter Awal'] = summary_df['Meter Awal'] -summary_df['Km Awal']*1000

            if 'Meter Akhir' in summary_df.columns:
                summary_df['Meter Akhir'] = summary_df['Meter Akhir']-summary_df['Km Akhir']*1000


            # Menentukan urutan kolom yang diinginkan
            desired_columns = (
                ['Track Number',
                 'Km Awal',
                  'Meter Awal',
                  'Km Akhir',
                    'Meter Akhir',
                      'Track Length'] +  # Kolom awal
                list(stdev_summary.keys()) +  # Kolom dari sheet_name (dinamis)
                ['TQI Total KAI','TQI Total EN']  # Kolom terakhir
            )

            # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df = summary_df[desired_columns]

            # Modifikasi index dengan mengalikan index dengan block_size
            summary_df.index = block_size / 4* (summary_df.index+1) 

            
            kota = []
            # Base data untuk Petak Jalan
            kota = databasekota(lines)
            # Fungsi untuk membuat base_data dari daftar kota
            def buat_base_data(kota):
                base_data = []
                for i in range(len(kota) - 1):
                    base_data.append({
                        "nama": f"{kota[i]['nama']}-{kota[i+1]['nama']}",
                        "first": kota[i]["posisi"],
                        "end": kota[i+1]["posisi"]
                    })
                return base_data

            # Membuat base_data
            base_data = buat_base_data(kota)

            # Menentukan `datakota` berdasarkan `trend`
            datakota = []
            for item in base_data:
                if trend == "up":
                    nama_split = item["nama"].split("-")
                    nama_balik = f"{nama_split[1]}-{nama_split[0]}"
                    datakota.append({
                        "nama": nama_balik,
                        "first": item["end"],
                        "end": item["first"]
                    })
                else:
                    datakota.append(item)


            

            # Menghitung awal dan akhir berdasarkan Km dan Meter
            awalPetakJalan = summary_df['Km Awal'] * 1000 + summary_df['Meter Awal']
            akhirPetakJalan = summary_df['Km Akhir'] * 1000 + summary_df['Meter Akhir']
            summary_df['NilaiTengah'] = (awalPetakJalan + akhirPetakJalan) / 2

            # Loop untuk memeriksa apakah titik stasiun ada di antara awal dan akhir
            for index, row in summary_df.iterrows():
                kota_terdeteksi = False  # Menandai apakah kota ditemukan pada baris ini
                
                for kota in datakota:
                    # Memastikan logika dinamis berdasarkan trend
                    lower_bound = min(kota["first"], kota["end"])
                    upper_bound = max(kota["first"], kota["end"])
                    
                    if lower_bound <= row['NilaiTengah'] <= upper_bound:
                        
                        # Menambahkan nama kota pada kolom 'Kota' di DataFrame summary_df
                        summary_df.loc[index, 'Petak Jalan'] = kota["nama"]
                        kota_terdeteksi = True
                        kotadilewati.append(kota["nama"])
                        break  # Jika kota ditemukan, hentikan pencarian lebih lanjut

                # Jika tidak ada kota yang terdeteksi, set "Tidak terdefinisi"
                if not kota_terdeteksi:
                    summary_df.loc[index, 'Petak Jalan'] = "Tidak terdefinisi"

            
            # Menghapus kolom 'NilaiTengah' setelah selesai digunakan
            summary_df = summary_df.drop(columns=['NilaiTengah'])

            # Menyimpan DataFrame StDevSummary ke sheet
            summary_df.to_excel(writer, index=False, header=True, sheet_name="StDevSummary")




            # Membuat salinan DataFrame summary_df
            # Membuat salinan dataframe untuk diproses
            summary_df_2 = summary_df.copy()

            # Menghapus kolom 'TQI Total EN' dari summary_df_2 jika ada
            if 'TQI Total EN' in summary_df_2.columns:
                summary_df_2.drop(columns=['TQI Total EN'], inplace=True)
                
            combined_columns = []
            def process_pairs(summary_df_2, pairs):
                # Daftar awalan untuk R dan L dengan prioritas
                r_prefixes = ["RIGHT", "Right", "right", "R", "r"]
                l_prefixes = ["LEFT", "Left", "left", "L", "l"]

                combined_columns = []

                for left_col, right_col in pairs:
                    # Cek apakah kedua kolom ada dalam dataframe
                    if left_col in summary_df_2.columns and right_col in summary_df_2.columns:
                        new_column_name = None

                        # Jika pasangan diawali dengan awalan "L" dan "R", deteksi dan buat nama kolom
                        for r_prefix in r_prefixes:
                            for l_prefix in l_prefixes:
                                # Jika kolom kiri dimulai dengan "L" dan kolom kanan dengan "R", atau sebaliknya
                                if left_col.startswith(l_prefix) and right_col.startswith(r_prefix):
                                    # Hapus prefix "L" atau "R" dari kolom kiri
                                    new_column_name = left_col[len(l_prefix):]
                                    break

                        # Jika tidak menggunakan awalan "L" dan "R", buat nama kolom berdasarkan pasangan lainnya
                        if not new_column_name:
                            new_column_name = f"{left_col}_{right_col}"  # Nama kolom berdasarkan pasangan yang lebih umum

                        # Tambahkan nama kolom baru ke dalam daftar combined_columns
                        combined_columns.append(new_column_name)

                        # Gabungkan kedua kolom dengan menghitung rata-rata dan beri nama kolom baru
                        summary_df_2[new_column_name] = (summary_df_2[left_col] + summary_df_2[right_col]) / 2

                        # Hapus kolom lama setelah rata-rata dihitung
                        summary_df_2.drop(columns=[left_col, right_col], inplace=True, errors='ignore')
                    else:
                        print(f"Kolom {left_col} atau {right_col} tidak ditemukan dalam dataframe.")  # Debugging jika kolom hilang

                return summary_df_2, combined_columns

            # Proses pasangan
            summary_df_2, combined_columns = process_pairs(summary_df_2, pairs)


            # Menentukan urutan kolom yang diinginkan
            desired_columns_2 = (
                ['Track Number','Km Awal','Meter Awal','Km Akhir', 'Meter Akhir', 'Track Length'] +  # Kolom awal
                combined_columns +  # Kolom yang sudah digabungkan berdasarkan pasangan
                remaining_params +  # Kolom yang tidak memiliki pasangan
                ['TQI Total KAI']  # Kolom terakhir
            )

            # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df_2 = summary_df_2[desired_columns_2]

            

            summary_df_2['Petak Jalan'] = summary_df['Petak Jalan']

            # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df_2 = summary_df_2.round(2)

            # Menyimpan DataFrame StDevSummary_2 ke sheet
            summary_df_2.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_2 (KAI)")




            # Membuat salinan DataFrame
            summary_df_3 = pd.DataFrame()
            endvalue = summary_df_2['Km Awal'].iloc[-1]  # Nilai terakhir dari kolom 'Km Awal'

            startpoint = kotadilewati[0].split('-')[0]
            endpoint = kotadilewati[-1].split('-')[1]


            # Menambahkan title manual sebelum header dan nilai-nilai
            title = [
                "DATA TRACK QUALITY INDEX (SESUAI PERDIR NOMOR : PER.U/KI.205/XII/1/KA-2023 )",
                f"TITIK AWAL: {startpoint}",
                f"KM AWAL: {origin_kilometer}",
                f"TITIK AKHIR: {endpoint}",
                f"KM AKHIR: {endvalue}",
                f"TRACK NUMBER: {lines}",
                f"ID: {idinput}",
                ""  # Baris kosong
            ]

            # Menulis DataFrame summary_df_3 ke dalam sheet "StDevSummary_3"
            summary_df_3.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_3 (KAI)")

            

            # Akses workbook dan sheet setelah menulis DataFrame
            workbook = writer.book
            sheet = workbook["StDevSummary_3 (KAI)"]

            sheet=titlecreate(sheet,title,summary_df_2,'center')
            
            sheet=wrapsheet(sheet,8,15,20)
            
            sheet=center_columns(sheet)

            sheet= font_set(workbook,sheet)


            

            for i in range(len(track_classes)):
                # Membuat DataFrame dengan 5 kolom kosong
                columns = [
                    'BATAS ATAS TRACK QUALITY (VERY GOOD)',
                    'BATAS ATAS TRACK QUALITY (GOOD)',
                    'BATAS ATAS TRACK QUALITY (FAIR)',
                    'BATAS ATAS TRACK QUALITY (POOR)',
                    'BATAS ATAS TRACK QUALITY (VERY POOR)'
                ]

                summary_temporary = pd.DataFrame(columns=columns)

                # Menghitung panjang segmen berdasarkan block_size, misal SAMPLES_PER_METER adalah 4 dan block_size adalah 160 maka segment_length adalah 40 meter berarti 1 block data stdev adalah 40 meter
                segment_length = segment_length_function(block_size, SAMPLES_PER_METER)
                # Menghitung nilai sementara berdasarkan tqi_kelas
                temporaryvalue = [tqi_kelas(tqi, i, segment_length) for tqi in summary_df['TQI Total KAI']]
                baris1=0
                baris2=0
                baris3=0
                baris4=0
                baris5=0
                for arrayunique in temporaryvalue:
                    baris1+=arrayunique[0]
                    baris2+=arrayunique[1]
                    baris3+=arrayunique[2]
                    baris4+=arrayunique[3]
                    baris5+=arrayunique[4]

                # Menyusun baris-baris data yang diperlukan
                rows = [
                    [np.nan] * 5,  # Baris kosong
                    [f"RESUME KELAS JALAN {track_classes[i]}","","","",""],
                    [
                        f'TOTAL PANJANG TRACK QUALITY (VERY GOOD) (M) TQI < {25+5*i}',
                        f'TOTAL PANJANG TRACK QUALITY (GOOD) (M) {25+5*i} <= TQI < {40+5*i}',
                        f'TOTAL PANJANG TRACK QUALITY (FAIR) (M) {40+5*i} <= TQI < {55+5*i}',
                        f'TOTAL PANJANG TRACK QUALITY (POOR) (M) {55+5*i} <= TQI < {70+5*i}',
                        f'TOTAL PANJANG TRACK QUALITY (VERY POOR) (M) {70+5*i} >= TQI',
                    ],
                    [baris1,baris2,baris3,baris4,baris5],
                    [np.nan] * 5,
                    [f"KELAS JALAN {track_classes[i]}","","","",""],
                    [
                    f'BATAS ATAS TRACK QUALITY (VERY GOOD) TQI < {25+5*i}',
                    f'BATAS ATAS TRACK QUALITY (GOOD) {25+5*i} <= TQI < {40+5*i}',
                    f'BATAS ATAS TRACK QUALITY (FAIR) {40+5*i} <= TQI < {55+5*i}',
                    f'BATAS ATAS TRACK QUALITY (POOR) {55+5*i} <= TQI < {70+5*i}',
                    f'BATAS ATAS TRACK QUALITY (VERY POOR) {70+5*i} >= TQI',
                    ],
                ]
                # Menggabungkan baris-baris menjadi satu DataFrame
                for row in rows:
                    summary_temporary = pd.concat([summary_temporary, pd.DataFrame([row], columns=columns)], ignore_index=True)

                

                # Menambahkan hasil perhitungan ke DataFrame
                summary_temporary = pd.concat([summary_temporary, pd.DataFrame(temporaryvalue, columns=columns)], ignore_index=True)

                
                # Menyimpan ke dalam file Excel
                summary_temporary.to_excel(
                    writer, 
                    index=False, 
                    header=False, 
                    startrow=2, 
                    sheet_name=f"KLASIFIKASI KELAS JALAN {track_classes[i]}"
                )

                
                # Akses workbook dan sheet setelah menulis DataFrame
                workbook = writer.book
                sheetkhusus = workbook[f"KLASIFIKASI KELAS JALAN {track_classes[i]}"]
                # Menambahkan title manual sebelum header dan nilai-nilai
                title_khusus = [
                    "KLASIFIKASI NILAI TQI  TERHADAP TRACK QUALITY & KELAS JALAN ",
                    f"SESUAI PERDIR NOMOR : PER.U/KI.205/XII/1/KA-2023 ",
                ]
                
                sheetkhusus=titlecreate(sheetkhusus,title_khusus,summary_temporary,'center')
                sheetkhusus=wrapsheet(sheetkhusus,5,15,30)
                sheetkhusus=center_columns(sheetkhusus)

                sheetkhusus= font_set(workbook,sheetkhusus)
                


                # Hapus baris ke-3
                # Hapus baris ke-3
                sheetkhusus.delete_rows(3)

                # Merge baris ke-4 pada kolom A hingga E
                sheetkhusus.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
                sheetkhusus.cell(row=4, column=1, value=f"RESUME KELAS JALAN {track_classes[i]}")

                # Atur alignment tengah (horizontal dan vertikal)
                sheetkhusus.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")

                # Merge baris ke-8 pada kolom A hingga E
                sheetkhusus.merge_cells(start_row=8, start_column=1, end_row=8, end_column=5)
                sheetkhusus.cell(row=8, column=1, value=f"KELAS JALAN {track_classes[i]}")

                # Atur alignment tengah untuk baris ke-8
                sheetkhusus.cell(row=8, column=1).alignment = Alignment(horizontal="center", vertical="center")


            if trend == "down":
                pass  # Tidak terjadi apa-apa
            elif trend == "up":
                datakota.reverse()  # Membalik urutan datakota
            
            
            for i in range(len(track_classes)):
                # Definisi kolom DataFrame
                # Definisi kolom DataFrame
                columns = [
                    'LINTAS', 'TRACK NUMBER', 'KM AWAL (METER)', 'KM AKHIR (METER)', 'PANJANG (METER)',

                    f'TOTAL PANJANG TRACK QUALITY (VERY GOOD) (M) TQI < {25+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (GOOD) (M) {25+5*i} <= TQI < {40+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (FAIR) (M) {40+5*i} <= TQI < {55+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (POOR) (M) {55+5*i} <= TQI < {70+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (VERY POOR) (M) {70+5*i} >= TQI',
                    'TQI'
                ]

                rows = []
                for kota in datakota:
                    # Memfilter DataFrame berdasarkan Petak Jalan
                    filtered_df = summary_df[summary_df['Petak Jalan'] == kota["nama"]]
                    
                    # Jika DataFrame kosong, lewati iterasi ini
                    if filtered_df.empty:
                        print(f"Petak Jalan untuk kota {kota['nama']} tidak ditemukan, melewati iterasi.")
                        continue

                    # Menghitung nilai sementara berdasarkan tqi_kelas
                    temporaryvalue = [tqi_kelas(tqi, i, block_size / 4) for tqi in filtered_df['TQI Total KAI']]

                    # Inisialisasi variabel akumulasi
                    baris1, baris2, baris3, baris4, baris5 = 0, 0, 0, 0, 0
                    for arrayunique in temporaryvalue:
                        baris1 += arrayunique[0]
                        baris2 += arrayunique[1]
                        baris3 += arrayunique[2]
                        baris4 += arrayunique[3]
                        baris5 += arrayunique[4]

                    # Perhitungan KM Awal dan KM Akhir
                    kmawal = int((filtered_df['Km Awal'].iloc[0] * 1000) + filtered_df['Meter Awal'].iloc[0])
                    kmakhir = int((filtered_df['Km Akhir'].iloc[-1] * 1000) + filtered_df['Meter Akhir'].iloc[-1])

                    # Total panjang dan rata-rata TQI
                    sumbaris = baris1 + baris2 + baris3 + baris4 + baris5
                    tqiakhir = sum(filtered_df['TQI Total KAI'])
                    tqiakhiraverage = tqiakhir / len(filtered_df['TQI Total KAI'])

                    # Menyusun baris data
                    rows.append([
                        kota["nama"], lines, kmawal, kmakhir, 
                        sumbaris, baris1, baris2, baris3, baris4, baris5,round(tqiakhiraverage, 2)
                    ])

                # Membuat DataFrame sementara
                summary_temporary_2 = pd.DataFrame(rows, columns=columns)

                # Menyimpan ke Excel
                summary_temporary_2.to_excel(
                    writer,
                    index=False,
                    header=True,
                    sheet_name=f"SUMMARY REPORT KELAS JALAN {track_classes[i]}"
                )

                

                # Akses workbook dan sheet untuk modifikasi lebih lanjut
                workbook = writer.book
                sheetkhusus_2 = workbook[f"SUMMARY REPORT KELAS JALAN {track_classes[i]}"]

                # Menambahkan judul manual sebelum header
                title_khusus_2 = ["SUMMARY REPORT",""]
                
                sheetkhusus_2 = center_columns(sheetkhusus_2)
                sheetkhusus_2 = titlecreate(sheetkhusus_2, title_khusus_2, summary_temporary_2,'center')
                sheetkhusus_2=wrapsheet(sheetkhusus_2,3,15,20)
                sheetkhusus_2=center_columns(sheetkhusus_2)

                sheetkhusus_2= font_set(workbook,sheetkhusus_2)

            summary_df_2_en = summary_df_2.copy()

            # Menghapus kolom 'TQI Total EN' dari summary_df_2 jika ada
            if 'TQI Total KAI' in summary_df_2_en.columns:
                summary_df_2_en.drop(columns=['TQI Total KAI'], inplace=True)
            
            summary_df_2_en['TQI Total EN'] = summary_df['TQI Total EN']

            # Menghapus kolom 'NilaiTengah' setelah selesai digunakan
            summary_df_2_en = summary_df_2_en.drop(columns=['Petak Jalan'])

            summary_df_2_en['Petak Jalan'] = summary_df['Petak Jalan']

            # Menyimpan ke Excel
            summary_df_2_en.to_excel(
                writer,
                index=False,
                header=True,
                sheet_name=f"StDevSummary_2 (EN)"
            )

            summary_df_3_en = summary_df_2_en.copy()
            
            # Menambahkan title manual sebelum header dan nilai-nilai
            title3= [
                "DATA TRACK QUALITY INDEX (SESUAI EN 13848-6 CoSD )",
                f"TITIK AWAL: {startpoint}",
                f"KM AWAL: {origin_kilometer}",
                f"TITIK AKHIR: {endpoint}",
                f"KM AKHIR: {endvalue}",
                f"TRACK NUMBER: {lines}",
                f"ID: {idinput}",
                ""  # Baris kosong
            ]

            # Menulis DataFrame summary_df_3 ke dalam sheet "StDevSummary_3"
            summary_df_3_en.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_3 (EN)")

            
            # Akses workbook dan sheet setelah menulis DataFrame
            workbook = writer.book
            sheet3 = workbook["StDevSummary_3 (EN)"]

            sheet3=titlecreate(sheet3,title3,summary_df_3_en,'center')
            sheet3=wrapsheet(sheet3,8,15,20)
            sheet3=center_columns(sheet3)
            sheet3= font_set(workbook,sheet3)

            

            # Menambahkan grafik ke dalam sheet StDevSummary
            for idx, column in enumerate(summary_df.columns):
                fig, ax = plt.subplots()
                ax.plot(summary_df.index, summary_df[column], label=column)
                ax.set_title(f"Grafik {column}")
                ax.set_xlabel('Km')
                ax.set_ylabel('TQI')
                ax.legend()

                # Simpan grafik ke dalam buffer
                img_stream = BytesIO()
                fig.savefig(img_stream, format='png')
                img_stream.seek(0)

                # Masukkan gambar ke dalam worksheet
                image = Image(img_stream)
                sheet = writer.sheets["StDevSummary"]
                image.anchor = f"A{len(summary_df) + idx + 2}"  # Mengatur posisi gambar di bawah tabel
                sheet.add_image(image)


        # Contoh penggunaan
        xlsx_path = 'allparameter.xlsx'  # Ganti dengan path file Excel Anda
        

        title= [
            "DATA TRACK QUALITY INDEX",
            "DATA TRACK QUALITY INDEX (SESUAI PERDIR NOMOR : PER.U/KI.205/XII/1/KA-2023 )",
            f"TITIK AWAL: {startpoint}",
            f"KM AWAL: {origin_kilometer}",
            f"TITIK AKHIR: {endpoint}",
            f"KM AKHIR: {endvalue}",
            f"TRACK NUMBER: {lines}",
            f"ID: {idinput}",
            ""  # Baris kosong
        ]
        
        sheet_name = "StDevSummary_3 (KAI)"  # Ganti dengan nama sheet yang ingin diekspor
        output_pdf_path = 'REPORT TQI KAI.pdf'  # Path untuk menyimpan PDF
        
        export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path,title,8,[], include_header=True)

        sheet_name = "StDevSummary_3 (EN)"  # Ganti dengan nama sheet yang ingin diekspor
        output_pdf_path = 'REPORT TQI EN.pdf'  # Path untuk menyimpan PDF

        title_2= [
            "DATA TRACK QUALITY INDEX",
            "DATA TRACK QUALITY INDEX (SESUAI EN 13848-6 CoSD )",
            f"TITIK AWAL: {startpoint}",
            f"KM AWAL: {origin_kilometer}",
            f"TITIK AKHIR: {endpoint}",
            f"KM AKHIR: {endvalue}",
            f"TRACK NUMBER: {lines}",
            f"ID: {idinput}",
            ""  # Baris kosong
        ]
        
        export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path,title_2,8,[], include_header=True)


        for i in range(len(track_classes)):
            sheet_name = f"SUMMARY REPORT KELAS JALAN {track_classes[i]}"
            output_pdf_path = f"SUMMARY REPORT KELAS JALAN {track_classes[i]}.pdf"
            title = [f"SUMMARY REPORT KELAS JALAN {track_classes[i]}"]
            export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path, title, 2, [], include_header=True)

            sheet_name = f"KLASIFIKASI KELAS JALAN {track_classes[i]}"
            output_pdf_path = f"KLASIFIKASI KELAS JALAN {track_classes[i]}.pdf"
            title = [
                f"KLASIFIKASI NILAI TQI TERHADAP TRACK QUALITY & KELAS JALAN",
                f"SESUAI PERDIR NOMOR : PER.U/KI.205/XII/1/KA-2023",
                f"Base Raw Data: {Path(file_path).name}",
                ]
            export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path, title, None, [[4, 7], [9, 'endrow']], include_header=True)

        # Hitung jumlah segmen per kategori kualitas untuk setiap kelas jalur
        segment_length = segment_length_function(block_size, SAMPLES_PER_METER)
        result = {cls: {'Very Good': 0, 'Good': 0, 'Fair': 0, 'Poor': 0, 'Very Poor': 0} for cls in track_classes}

        for i, cls in enumerate(track_classes):
            # Klasifikasikan setiap nilai TQI
            for tqi in summary_df['TQI Total KAI']:
                classification = tqi_kelas(tqi, i, segment_length)
                # classification adalah array [v_good, good, fair, poor, v_poor]
                if classification[0] > 0:
                    result[cls]['Very Good'] += 1
                elif classification[1] > 0:
                    result[cls]['Good'] += 1
                elif classification[2] > 0:
                    result[cls]['Fair'] += 1
                elif classification[3] > 0:
                    result[cls]['Poor'] += 1
                elif classification[4] > 0:
                    result[cls]['Very Poor'] += 1

        # Format hasil untuk ditampilkan di GUI
        result_str = "\n".join(
            f"Kelas {cls}: Very Good={counts['Very Good']}, Good={counts['Good']}, "
            f"Fair={counts['Fair']}, Poor={counts['Poor']}, Very Poor={counts['Very Poor']}"
            for cls, counts in result.items()
        )

        # Kembalikan hasil
        return result_str

    except Exception as e:
        return f"Terjadi kesalahan: {str(e)}"

# Layout GUI
layout = [
    [
        sg.Column([  
            [sg.Text("-----------PERHATIAN-----------")],

            [sg.Text("Gunakan awalan LEFT dan RIGHT jika ada parameter yang berpasangan.")],
            [sg.Text("Contoh: gunakan LEFTvelocity1 dan RIGHTvelocity1.")],
            [sg.Text("Sistem akan otomatis mengenali dan memproses pasangan parameter tersebut.")],
            [sg.Text("Pilih file CV:")],
            [sg.Input(key="-FILE-"), sg.FileBrowse(file_types=(("CSV Files", "*.csv"),))],
            [sg.Button("Preview Header")],
            [sg.Text("Pilih parameter yang ingin diambil:")],
            [sg.Listbox(values=[], select_mode="multiple", size=(50, 10), key="-HEADERS-")],

            [sg.Text("Pilih Meter (Misal pilih 40 maka tiap 40 meter 1 data TQI):")],
            [sg.Combo(meter_capture_values, default_value="40", key="-meter_capture_value-")],
            
        ], vertical_alignment='top'),  
        sg.VerticalSeparator(),  
        sg.Column([
            # [sg.Text("-----------PERHATIAN-----------")],

            # [sg.Text("Nilai 0 tidak akan dihitung sehingga dianggap NaN.")],
            # [sg.Text("Nilai NaN akan diabaikan dalam perhitungan.")],
            # [sg.Text("Sistem akan melakukan drop terhadap nilai NaN dimana jumlah sampel yang digunakan pasti lebih sedikit.")],

            [sg.Text("ID")],
            [sg.Input(size=(10,1), key="-ID_FIRST-"), sg.Text("-"), sg.Input(size=(10,1), key="-ID_END-")],
            [sg.Text("Nilai Awal (Km) (Penggunaan . diperbolehkan dan kosong untuk dipilih otomatis)")],
            [sg.Input(key="-origin_kilometer-")],
            [sg.Text("Pasangan Parameter:")],
            [sg.Listbox(values=[], select_mode="multiple", size=(50, 10), key="-PAIRS-")],
            [sg.Button("Proses"), sg.Button("Keluar")],
            [sg.Text("", size=(50, 5), key="-OUTPUT-")]
        ], vertical_alignment='top')  
    ]
]

# Membuat jendela GUI
window = sg.Window("TQI Calculation Report", layout)

# Event loop
while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED or event == "Keluar":
        break

    if event == "Preview Header":
        file_path = values["-FILE-"]
        if not file_path:
            window["-OUTPUT-"].update("Harap pilih file CSV terlebih dahulu.")
        else:
            try:
                def detect_encoding(file_path):
                    # Membaca file dalam mode byte untuk mendeteksi encoding
                    with open(file_path, 'rb') as f:
                        raw_data = f.read()
                        result = chardet.detect(raw_data)
                    return result['encoding']
                encoding = detect_encoding(file_path)

                data = pd.read_csv(file_path , sep=',', encoding=encoding, skiprows=1)

                headers = data.columns.tolist()
                window["-HEADERS-"].update(headers)
                window["-OUTPUT-"].update("Header berhasil dimuat. Pilih parameter yang diinginkan.")
            except Exception as e:
                window["-OUTPUT-"].update(f"Terjadi kesalahan saat membaca file: {e}")

    if event == "Proses":
        file_path = values["-FILE-"]
        selected_headers = values["-HEADERS-"]
        selected_pairs = values["-PAIRS-"]
        
        # Generate pasangan parameter berdasarkan header yang dipilih
        if selected_headers:
            pairs = generate_pairs(selected_headers)
            pair_strings = [f'{pair[0]} - {pair[1]}' for pair in pairs]
            window["-PAIRS-"].update(pair_strings)
        
        #block_size (Ukuran Blok)
        #block_size menentukan berapa banyak baris data yang dimasukkan ke dalam satu blok (segmen) untuk menghitung standar deviasi (stdev) setiap parameter.
        #Misalnya, jika block_size = 160, setiap blok berisi 160 baris data, yang mewakili satu segmen jalur (40 meter).

        block_size = int(values["-meter_capture_value-"]) * 4
        
        # "Origin" menunjukkan titik asal pengukuran dalam kilometer
        origin_kilometer = values["-origin_kilometer-"]
        

        idinput = f"{values['-ID_FIRST-']}-{values['-ID_END-']}"


        

        if not file_path or not selected_headers:
            window["-OUTPUT-"].update("Harap masukkan file dan pilih parameter yang valid.")
        else:
            pairs = []
            for pair_str in selected_pairs:
                header1, header2 = pair_str.split(' - ')
                pairs.append((header1, header2))

            result = process_csv(file_path, selected_headers, block_size, origin_kilometer,encoding,idinput)
            window["-OUTPUT-"].update(result)

window.close()
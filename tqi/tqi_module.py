import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl.drawing.image import Image      
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from fpdf import FPDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Fungsi untuk mengkategorikan Track Quality Index (TQI) berdasarkan kelas
def tqi_kelas(tqi,kelas,inputer):
    if(kelas==0):
        if tqi<(25):
            return [inputer,0,0,0,0] #BATAS ATAS TRACK QUALITY  (VERY GOOD)
        elif (35)>=tqi>(25):
            return [0,inputer,0,0,0] #BATAS ATAS TRACK QUALITY  (GOOD)
        elif (50)>=tqi>(35):
            return [0,0,inputer,0,0] #BATAS ATAS TRACK QUALITY  (FAIR)
        elif (65)>=tqi>(50):
            return [0,0,0,inputer,0] #BATAS ATAS TRACK QUALITY  (POOR)
        else:
            return [0,0,0,0,inputer] #BATAS ATAS TRACK QUALITY  (VERY POOR)
    elif(kelas==1):
        if tqi<(30):
            return [inputer,0,0,0,0] #BATAS ATAS TRACK QUALITY  (VERY GOOD)
        elif (45)>=tqi>(30):
            return [0,inputer,0,0,0] #BATAS ATAS TRACK QUALITY  (GOOD)
        elif (60)>=tqi>(45):
            return [0,0,inputer,0,0] #BATAS ATAS TRACK QUALITY  (FAIR)
        elif (75)>=tqi>(60):
            return [0,0,0,inputer,0] #BATAS ATAS TRACK QUALITY  (POOR)
        else:
            return [0,0,0,0,inputer] #BATAS ATAS TRACK QUALITY  (VERY POOR)
    elif(kelas==2):
        if tqi<(40):
            return [inputer,0,0,0,0] #BATAS ATAS TRACK QUALITY  (VERY GOOD)
        elif (55)>=tqi>(40):
            return [0,inputer,0,0,0] #BATAS ATAS TRACK QUALITY  (GOOD)
        elif (70)>=tqi>(55):
            return [0,0,inputer,0,0] #BATAS ATAS TRACK QUALITY  (FAIR)
        elif (85)>=tqi>(70):
            return [0,0,0,inputer,0] #BATAS ATAS TRACK QUALITY  (POOR)
        else:
            return [0,0,0,0,inputer] #BATAS ATAS TRACK QUALITY  (VERY POOR)
    elif(kelas==3):
        if tqi<(50):
            return [inputer,0,0,0,0] #BATAS ATAS TRACK QUALITY  (VERY GOOD)
        elif (65)>=tqi>(50):
            return [0,inputer,0,0,0] #BATAS ATAS TRACK QUALITY  (GOOD)
        elif (80)>=tqi>(65):
            return [0,0,inputer,0,0] #BATAS ATAS TRACK QUALITY  (FAIR)
        elif (95)>=tqi>(80):
            return [0,0,0,inputer,0] #BATAS ATAS TRACK QUALITY  (POOR)
        else:
            return [0,0,0,0,inputer] #BATAS ATAS TRACK QUALITY  (VERY POOR)
    
# Fungsi untuk mengekspor file Excel ke PDF
def export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path, title, skiprows=None, startend=None, include_header=True):
    # startend berbasis 1 bukan 0
    # Baca file Excel dan isi merged cells
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, skiprows=skiprows, header=None)
    df = df.ffill(axis=0)  # Mengisi merged cells ke bawah
    df = df.fillna("")  # Mengisi sel kosong dengan string kosong

    # Inisialisasi PDF
    pdf = SimpleDocTemplate(output_pdf_path, pagesize=landscape(A4))
    pdf.title = " | ".join(title)  # Menambahkan setTitle()
    elements = []

    # Gaya teks dan tabel
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', parent=styles['Title'], fontSize=12, spaceAfter=12, alignment=0)
    cell_style = ParagraphStyle(name='Cell', parent=styles['BodyText'], fontSize=8, leading=10, alignment=1)

    table_style = TableStyle([ 
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    # Tambahkan judul ke PDF
    for line in title:
        elements.append(Paragraph(line, title_style))
    elements.append(Spacer(1, 12))

    # Fungsi untuk membungkus teks dalam tabel
    def wrap_text(text):
        return Paragraph(str(text) if pd.notna(text) else "", cell_style)

    # Handle start dan end jika ada
    if startend:
        for start, end in startend:
            # Sesuaikan indeks berbasis 1 ke berbasis 0
            start_index = start - 1
            end_index = None if end == 'endrow' else end - 1

            # Ambil subset data
            subset = df.iloc[start_index:end_index]

            # Ambil semua data dengan wrap
            alldata = [[wrap_text(cell) for cell in row] for row in subset.values.tolist()]

            # Pisahkan header dan value dari alldata
            if include_header:
                header = alldata[0]  # Baris pertama sebagai header
                values = alldata[1:]  # Sisanya sebagai values
            else:
                header = []
                values = alldata

            # Gabungkan header dan values jika diperlukan
            if include_header:
                data = [header] + values
            else:
                data = values


            # Membuat tabel
            table = Table(data, repeatRows=1 if include_header else 0)
            table.setStyle(table_style)
            elements.append(table)
            elements.append(Spacer(1, 20))
            elements.append(PageBreak())
    else:
        # Jika tidak ada startend, ambil data seluruhnya
        alldata = [[wrap_text(cell) for cell in row] for row in df.values.tolist()]

        # Pisahkan header dan value dari alldata
        if include_header:
            header = alldata[0]  # Baris pertama sebagai header
            values = alldata[1:]  # Sisanya sebagai values
        else:
            header = []
            values = alldata
            
        # Gabungkan header dan values jika diperlukan
        if include_header:
            data = [header] + values
        else:
            data = values


        # Membuat tabel
        table = Table(data, repeatRows=1 if include_header else 0)
            
        table.setStyle(table_style)
        elements.append(table)

    # Build PDF
    pdf.build(elements)

# Fungsi untuk menghasilkan pasangan header yang sesuai
def generate_pairs(selected_headers):
    pairs = set()  # Gunakan set untuk memastikan pasangan unik
    r_prefixes = ["RIGHT", "Right", "right", "R", "r"]
    l_prefixes = ["LEFT", "Left", "left", "L", "l"]
    
    for header in selected_headers:
        # Cek jika header dimulai dengan salah satu awalan di r_prefixes
        for r_prefix in r_prefixes:
            if header.startswith(r_prefix):
                for l_prefix in l_prefixes:
                    pair = header.replace(r_prefix, l_prefix, 1)
                    if pair in selected_headers:
                        # Urutkan pasangan sebelum dimasukkan ke dalam set
                        pairs.add(tuple(sorted((header, pair))))
                        break
                break
        
        # Cek jika header dimulai dengan salah satu awalan di l_prefixes
        for l_prefix in l_prefixes:
            if header.startswith(l_prefix):
                for r_prefix in r_prefixes:
                    pair = header.replace(l_prefix, r_prefix, 1)
                    if pair in selected_headers:
                        # Urutkan pasangan sebelum dimasukkan ke dalam set
                        pairs.add(tuple(sorted((header, pair))))
                        break
                break
    
    # Ubah set ke list agar lebih mudah digunakan
    return list(pairs)


# Fungsi untuk validasi baris yang valid
def get_valid_block_range(file_path, selected_headers, encoding='utf-8'):
    # Baca data dengan skip baris header tambahan
    temp_data = pd.read_csv(file_path, sep=',', encoding=encoding, skiprows=1)

    if not set(selected_headers).issubset(temp_data.columns):
        raise ValueError("Beberapa header yang dipilih tidak ada di file CSV.")

    selected_columns = temp_data[selected_headers].copy()

    # Ubah 0 dan 0.00 menjadi NaN
    for col in selected_columns.columns:
        if selected_columns[col].dtype in ['float64', 'int64']:
            selected_columns[col] = selected_columns[col].replace([0, 0.00], np.nan)

    # Deteksi baris yang mengandung 'NV'
    invalid_nv = selected_columns.apply(
        lambda x: x.astype(str).str.contains('NV', na=False), axis=1
    ).any(axis=1)

    # Mask baris valid
    valid_mask = (~invalid_nv) & selected_columns.notna().all(axis=1)

    if not valid_mask.any():
        return None, 0

    start_row = valid_mask[valid_mask].index[0]
    end_row = valid_mask[valid_mask].index[-1]
    rows_to_read = end_row - start_row + 1

    return start_row, rows_to_read

# Fungsi untuk mengatur alignment kolom menjadi center
def center_columns(sheet):
    # Iterasi melalui setiap kolom di sheet
    for col in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            # Cek apakah wrap_text sudah True
            current_alignment = cell.alignment
            wrap_status = current_alignment.wrap_text if current_alignment else False
            # Set alignment center dan pertahankan wrap_text jika sudah ada
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center', 
                wrap_text=wrap_status
            )


    return sheet

# Fungsi untuk membungkus teks di dalam sel dan menyesuaikan tinggi baris dan lebar kolom
def wrapsheet(sheet, start_row=1, width_value=15, height_value=20):
    # Penyesuaian tinggi baris dan lebar kolom
    for row in range(start_row, sheet.max_row + 1):
        max_row_height = 1
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell_length = len(str(cell.value))
                estimated_lines = (cell_length // width_value) + 1
                max_row_height = max(max_row_height, estimated_lines)
        
        sheet.row_dimensions[row].height = max_row_height * height_value

    # Penyesuaian lebar kolom
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, width_value)
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Lakukan wrap text dengan pengecekan alignment
    for row in range(start_row, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                if cell.alignment.horizontal == 'center' and cell.alignment.vertical == 'center':
                    # Jika sudah center, tambahkan wrap_text dengan mempertahankan center alignment
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                else:
                    # Jika tidak, hanya set wrap_text
                    cell.alignment = Alignment(wrap_text=True)

    return sheet

# Fungsi untuk membuat judul pada sheet
def titlecreate(sheet, title, summary_df_2,position):
    if position=="normal":
        # Bersihkan seluruh sheet
        sheet.delete_rows(1, sheet.max_row)

        # Menulis title
        for idx, value in enumerate(title):
            sheet.cell(row=idx + 1, column=1, value=value)
            sheet.cell(row=idx + 1, column=1).alignment = Alignment(horizontal="center")  # Alignment tengah

        # Menulis header summary_df_2 di bawah title
        header_row = len(title) + 1
        for col_idx, header in enumerate(summary_df_2.columns, start=1):
            sheet.cell(row=header_row, column=col_idx, value=header)
            sheet.cell(row=header_row, column=col_idx).alignment = Alignment(horizontal="center")

        # Menulis data summary_df_2 di bawah header
        for r_idx, row in enumerate(summary_df_2.itertuples(index=False), start=header_row + 1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value if not pd.isnull(value) else "")

        # Menyesuaikan lebar kolom
        for col_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            col_letter = col_cells[0].column_letter
            sheet.column_dimensions[col_letter].width = max_length + 2
        return sheet
    elif position=="center":
        # Bersihkan seluruh sheet
        sheet.delete_rows(1, sheet.max_row)

        # Menulis title dan melakukan merge
        max_col = len(summary_df_2.columns)
        for idx, value in enumerate(title):
            merge_start = 1
            merge_end = max_col if max_col > 1 else 1
            sheet.merge_cells(start_row=idx + 1, start_column=merge_start, end_row=idx + 1, end_column=merge_end)
            cell = sheet.cell(row=idx + 1, column=1, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Menulis header summary_df_2 di bawah title
        header_row = len(title) + 1
        for col_idx, header in enumerate(summary_df_2.columns, start=1):
            sheet.cell(row=header_row, column=col_idx, value=header)
            sheet.cell(row=header_row, column=col_idx).alignment = Alignment(horizontal="center")

        # Menulis data summary_df_2 di bawah header
        for r_idx, row in enumerate(summary_df_2.itertuples(index=False), start=header_row + 1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value if not pd.isnull(value) else "")

        # Menyesuaikan lebar kolom (dipersempit agar lebih rapi)
        for col_idx, col_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            max_length = max_length if max_length > 10 else 10  # Minimum width 10
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = max_length + 1  # Dipersempit dengan menambahkan 1 saja

        return sheet

# Fungsi untuk mengubah nama sheet dengan menambahkan "SD_" di depan atau setelah prefix
def transform_name(sheet_name):
    r_prefixes = ["RIGHT", "Right", "right"]
    l_prefixes = ["LEFT", "Left", "left"]
    
    # Jika nama sheet diawali dengan "RIGHT" atau "LEFT"
    for prefix in r_prefixes + l_prefixes:
        if sheet_name.startswith(prefix):
            # Jika diawali dengan prefix, tambahkan "SD_" setelah prefix
            return sheet_name.replace(prefix, prefix + "SD_", 1)
    
    # Jika tidak ada prefix, tambahkan "SD_" di depan
    return "SD_" + sheet_name

# Fungsi untuk mengatur font pada workbook dan sheet
def font_set(workbook, sheet):
    # Create the style if it doesn't exist
    # arial_style = NamedStyle(name="arial_style")
    # arial_style.font = Font(name='Arial', size=11)

    # # Only add the style if it does not already exist in the workbook
    # if arial_style.name not in workbook.named_styles:
    #     workbook.add_named_style(arial_style)

    # # Apply the style to each cell in the sheet
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         cell.style = arial_style

    return sheet

# Fungsi untuk menentukan arah perubahan nilai
def upordown(before,after):
    if before > after:
        return "down"
    elif before < after:
        return "up"
    else:
        return "same"  # Jika nilai sebelum dan sesudah sama

# Fungsi untuk mendapatkan data kota berdasarkan jalur
def databasekota(lines):
    kota = []
    if(lines=='Line 1'):
        # Base data untuk Petak Jalan
        kota = [
            {"nama": "Harjamukti", "posisi": 14524},
            {"nama": "Ciracas", "posisi": 8864},
            {"nama": "Kampung Rambutan", "posisi": 7262},
            {"nama": "TMII", "posisi": 5315},
            {"nama": "TITIK 0", "posisi": 0},
        ]
    elif(lines=='Line 2'):
        # Base data untuk Petak Jalan
        kota = [
            {"nama": "TITIK 0", "posisi": 0},
             {"nama": "Cawang", "posisi": -187},
             {"nama": "Ciliwung", "posisi": -1445},
             {"nama": "Cikoko", "posisi": -2221},
             {"nama": "Pancoran", "posisi": -4305},
            {"nama": "Kuningan", "posisi": -6868},
             {"nama": "Rasuna Said", "posisi": -7666},
              {"nama": "Setiabudi", "posisi": -9062},
              {"nama": "Dukuh Atas", "posisi": -9853},
        ]
    elif(lines=='Line 3'):
        # Base data untuk Petak Jalan
        kota = [
            {"nama": "Jati Mulya", "posisi": 17408},
            {"nama": "Bekasi Barat", "posisi": 13702},
            {"nama": "Cikunir 2", "posisi": 10441},
            {"nama": "Cikunir 1", "posisi": 9172},
            {"nama": "Jatibening Baru", "posisi": 6508},
            {"nama": "Halim", "posisi": 1367},
            {"nama": "TITIK 0", "posisi": 0},
        ]
    return kota


# Fungsi untuk menghitung panjang segmen berdasarkan block_size dan SAMPLES_PER_METER
# Jika block_size adalah 160 dan sample_per_meter 4 maka panjang segmen adalah 40 meter.
# Dimana block_size adalah jumlah baris data yang dimasukkan ke dalam satu blok (segmen) untuk menghitung standar deviasi (stdev) setiap parameter.
# Dimana SAMPLES_PER_METER adalah jumlah sampel per meter yang didefinisikan misal nilainya 4 berarti setiap 4 sampel adalah 1 meter.

def segment_length_function(block_size, SAMPLES_PER_METER):
    """
    Fungsi untuk menghitung panjang segmen berdasarkan block_size dan SAMPLES_PER_METER.
    """
    if block_size <= 0 or SAMPLES_PER_METER <= 0:
        raise ValueError("block_size dan SAMPLES_PER_METER harus bilangan positif.")
    
    segment_length = block_size / SAMPLES_PER_METER
    return segment_length #satuannya adalah meter 


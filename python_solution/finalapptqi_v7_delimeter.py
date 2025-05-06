import pandas as pd
import PySimpleGUI as sg
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image
import re
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from fpdf import FPDF
import chardet
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle





#### Kodingan ini berhasil untuk sistem custom untuk data mas ibnu dan sedang mengalami penyesuaian agar bisa merespon
#### data dari pak ndaru yang baru
#### Fitur preview dan choose, otomatis logic


##update V_2 -> menangani jika inputan csv

def tqi_kelas(tqi,kelas,inputer):
    if(kelas==0):
        if tqi<(25):
            return [inputer,0,0,0,0] #BATAS ATAS TRACK QUALITY  (VERY GOOD)
        elif (35)>=tqi>(25):
            return [0,inputer,0,0,0] #BATAS ATAS TRACK QUALITY  (GOOD)
        elif (50)>=tqi>(35):
            return [0,0,inputer,0,0] #BATAS ATAS TRACK QUALITY  (FAIR)
        elif (65)>=tqi>(50):
            return [0,0,0,inputer,0] #BATAS ATAS TRACK QUALITY  (POOR)
        else:
            return [0,0,0,0,inputer] #BATAS ATAS TRACK QUALITY  (VERY POOR)
    elif(kelas==1):
        if tqi<(30):
            return [inputer,0,0,0,0] #BATAS ATAS TRACK QUALITY  (VERY GOOD)
        elif (45)>=tqi>(30):
            return [0,inputer,0,0,0] #BATAS ATAS TRACK QUALITY  (GOOD)
        elif (60)>=tqi>(45):
            return [0,0,inputer,0,0] #BATAS ATAS TRACK QUALITY  (FAIR)
        elif (75)>=tqi>(60):
            return [0,0,0,inputer,0] #BATAS ATAS TRACK QUALITY  (POOR)
        else:
            return [0,0,0,0,inputer] #BATAS ATAS TRACK QUALITY  (VERY POOR)
    elif(kelas==2):
        if tqi<(40):
            return [inputer,0,0,0,0] #BATAS ATAS TRACK QUALITY  (VERY GOOD)
        elif (55)>=tqi>(40):
            return [0,inputer,0,0,0] #BATAS ATAS TRACK QUALITY  (GOOD)
        elif (70)>=tqi>(55):
            return [0,0,inputer,0,0] #BATAS ATAS TRACK QUALITY  (FAIR)
        elif (85)>=tqi>(70):
            return [0,0,0,inputer,0] #BATAS ATAS TRACK QUALITY  (POOR)
        else:
            return [0,0,0,0,inputer] #BATAS ATAS TRACK QUALITY  (VERY POOR)
    elif(kelas==3):
        if tqi<(50):
            return [inputer,0,0,0,0] #BATAS ATAS TRACK QUALITY  (VERY GOOD)
        elif (65)>=tqi>(50):
            return [0,inputer,0,0,0] #BATAS ATAS TRACK QUALITY  (GOOD)
        elif (80)>=tqi>(65):
            return [0,0,inputer,0,0] #BATAS ATAS TRACK QUALITY  (FAIR)
        elif (95)>=tqi>(80):
            return [0,0,0,inputer,0] #BATAS ATAS TRACK QUALITY  (POOR)
        else:
            return [0,0,0,0,inputer] #BATAS ATAS TRACK QUALITY  (VERY POOR)
    
# Fungsi untuk membaca sheet dan mengekspor ke PDF
def export_xlsx_to_pdf_fpdf(xlsx_path, sheet_name, output_pdf_path, title, skiprows, startend):
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, skiprows=skiprows) if skiprows is not None else pd.read_excel(xlsx_path, sheet_name=sheet_name)
    
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font('Arial', 'B', 8)

    # Tambahkan judul
    for line in title:
        pdf.cell(0, 8, txt=line, ln=True, align='L')
    
    pdf.ln(8)

    # Hitung lebar kolom dengan memperhatikan panjang teks maksimal di setiap kolom
    max_col_widths = []
    for col in df.columns:
        max_width = pdf.get_string_width(str(col)) + 6  # Tambahkan padding
        for value in df[col]:
            max_width = max(max_width, pdf.get_string_width(str(value)) + 6)
        max_col_widths.append(max_width)

    if skiprows is None and startend:
        # Jika menggunakan startend, abaikan header
        for start, end in startend:
            if end == 'endrow':
                subset = df.iloc[start - 1:]
            else:
                subset = df.iloc[start - 1:end]

            # Tambahkan spasi atau garis pemisah untuk memulai tabel baru
            pdf.ln(10)  # Jarak antar tabel
            
            # Isi data tabel tanpa header
            pdf.set_font('Arial', '', 6)
            for row in subset.itertuples(index=False):
                for idx, value in enumerate(row):
                    pdf.cell(max_col_widths[idx], 6, str(value), border=1, align='C')
                pdf.ln()

    else:
        # Jika skiprows terisi atau startend kosong, tambahkan header dan isi data
        for col, width in zip(df.columns, max_col_widths):
            pdf.cell(width, 6, col, border=1, align='C')
        pdf.ln()

        for row in df.itertuples(index=False):
            for idx, value in enumerate(row):
                pdf.cell(max_col_widths[idx], 6, str(value), border=1, align='C')
            pdf.ln()

    def footer():
        pdf.set_y(-15)
        pdf.set_font('Arial', 'I', 6)
        pdf.cell(0, 10, f'Page {pdf.page_no()}', 0, 0, 'R')

    pdf.footer = footer
    pdf.output(output_pdf_path)

import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path, title, skiprows=None, startend=None, include_header=True):
    # startend berbasis 1 bukan 0
    # Baca file Excel dan isi merged cells
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, skiprows=skiprows, header=None)
    df = df.ffill(axis=0)  # Mengisi merged cells ke bawah
    df = df.fillna("")  # Mengisi sel kosong dengan string kosong

    # Inisialisasi PDF
    pdf = SimpleDocTemplate(output_pdf_path, pagesize=landscape(A4))
    elements = []

    # Gaya teks dan tabel
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', parent=styles['Title'], fontSize=12, spaceAfter=12, alignment=0)
    cell_style = ParagraphStyle(name='Cell', parent=styles['BodyText'], fontSize=8, leading=10, alignment=1)

    table_style = TableStyle([ 
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    # Tambahkan judul ke PDF
    for line in title:
        elements.append(Paragraph(line, title_style))
    elements.append(Spacer(1, 12))

    # Fungsi untuk membungkus teks dalam tabel
    def wrap_text(text):
        return Paragraph(str(text) if pd.notna(text) else "", cell_style)

    # Handle start dan end jika ada
    if startend:
        for start, end in startend:
            # Sesuaikan indeks berbasis 1 ke berbasis 0
            start_index = start - 1
            end_index = None if end == 'endrow' else end - 1

            # Ambil subset data
            subset = df.iloc[start_index:end_index]

            # Ambil semua data dengan wrap
            alldata = [[wrap_text(cell) for cell in row] for row in subset.values.tolist()]

            # Pisahkan header dan value dari alldata
            if include_header:
                header = alldata[0]  # Baris pertama sebagai header
                values = alldata[1:]  # Sisanya sebagai values
            else:
                header = []
                values = alldata

            # Gabungkan header dan values jika diperlukan
            if include_header:
                data = [header] + values
            else:
                data = values


            # Membuat tabel
            table = Table(data, repeatRows=1 if include_header else 0)
            table.setStyle(table_style)
            elements.append(table)
            elements.append(Spacer(1, 20))
            elements.append(PageBreak())
    else:
        # Jika tidak ada startend, ambil data seluruhnya
        alldata = [[wrap_text(cell) for cell in row] for row in df.values.tolist()]

        # Pisahkan header dan value dari alldata
        if include_header:
            header = alldata[0]  # Baris pertama sebagai header
            values = alldata[1:]  # Sisanya sebagai values
        else:
            header = []
            values = alldata
            
        # Gabungkan header dan values jika diperlukan
        if include_header:
            data = [header] + values
        else:
            data = values


        # Membuat tabel
        table = Table(data, repeatRows=1 if include_header else 0)
            
        table.setStyle(table_style)
        elements.append(table)

    # Build PDF
    pdf.build(elements)


def generate_pairs(selected_headers):
    pairs = set()  # Gunakan set untuk memastikan pasangan unik
    r_prefixes = ["RIGHT", "Right", "right", "R", "r"]
    l_prefixes = ["LEFT", "Left", "left", "L", "l"]
    
    for header in selected_headers:
        # Cek jika header dimulai dengan salah satu awalan di r_prefixes
        for r_prefix in r_prefixes:
            if header.startswith(r_prefix):
                for l_prefix in l_prefixes:
                    pair = header.replace(r_prefix, l_prefix, 1)
                    if pair in selected_headers:
                        # Urutkan pasangan sebelum dimasukkan ke dalam set
                        pairs.add(tuple(sorted((header, pair))))
                        break
                break
        
        # Cek jika header dimulai dengan salah satu awalan di l_prefixes
        for l_prefix in l_prefixes:
            if header.startswith(l_prefix):
                for r_prefix in r_prefixes:
                    pair = header.replace(l_prefix, r_prefix, 1)
                    if pair in selected_headers:
                        # Urutkan pasangan sebelum dimasukkan ke dalam set
                        pairs.add(tuple(sorted((header, pair))))
                        break
                break
    
    # Ubah set ke list agar lebih mudah digunakan
    return list(pairs)

# Fungsi untuk validasi baris yang valid
def automatevalidrow(file_path, selected_headers, encoding):
    print(encoding)
    temp_data = pd.read_csv(file_path, sep=',', encoding=encoding, skiprows=1)

    # Validasi header yang dipilih
    if not set(selected_headers).issubset(temp_data.columns):
        raise ValueError("Beberapa header yang dipilih tidak ada di file CSV.")

    # Filter kolom berdasarkan header yang dipilih
    selected_columns = temp_data[selected_headers]

    # Buang baris yang memiliki nilai 'NV' atau '0.00' di kolom yang dipilih
    selected_columns = selected_columns[~selected_columns.apply(
        lambda x: x.astype(str).str.contains('NV|0\.00')
    ).any(axis=1)]

    # Identifikasi baris dengan angka valid di semua kolom yang dipilih
    numeric_rows = selected_columns.applymap(
        lambda x: isinstance(x, (int, float)) and x != 0.00 or
                  (isinstance(x, str) and x.replace('.', '', 1).isdigit() and x != '0.00')
    ).all(axis=1)

    start_row = numeric_rows[numeric_rows == False].index[0] if not numeric_rows.all() else None
    print(f"Start row: {start_row}")

    if start_row is None:
        return None, 0  # Jika semua baris valid, tidak perlu lanjutkan

    # Mencari end_row (baris terakhir yang False)
    end_row = numeric_rows[numeric_rows == False].index[-1]  # Baris terakhir yang False
    print(f"End row: {end_row}")

    # Menghitung jumlah baris yang akan dibaca
    rows_to_read = end_row - start_row + 1
    print(f"Rows to read: {rows_to_read}")

    return start_row, rows_to_read

def center_columns(sheet):
    # Iterasi melalui setiap kolom di sheet
    for col in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            # Cek apakah wrap_text sudah True
            current_alignment = cell.alignment
            wrap_status = current_alignment.wrap_text if current_alignment else False
            # Set alignment center dan pertahankan wrap_text jika sudah ada
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center', 
                wrap_text=wrap_status
            )


    return sheet

def wrapsheet(sheet, start_row=1, width_value=15, height_value=20):
    # Penyesuaian tinggi baris dan lebar kolom
    for row in range(start_row, sheet.max_row + 1):
        max_row_height = 1
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell_length = len(str(cell.value))
                estimated_lines = (cell_length // width_value) + 1
                max_row_height = max(max_row_height, estimated_lines)
        
        sheet.row_dimensions[row].height = max_row_height * height_value

    # Penyesuaian lebar kolom
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, width_value)
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Lakukan wrap text dengan pengecekan alignment
    for row in range(start_row, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                if cell.alignment.horizontal == 'center' and cell.alignment.vertical == 'center':
                    # Jika sudah center, tambahkan wrap_text dengan mempertahankan center alignment
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                else:
                    # Jika tidak, hanya set wrap_text
                    cell.alignment = Alignment(wrap_text=True)

    return sheet

def titlecreate(sheet, title, summary_df_2,position):
    if position=="normal":
        # Bersihkan seluruh sheet
        sheet.delete_rows(1, sheet.max_row)

        # Menulis title
        for idx, value in enumerate(title):
            sheet.cell(row=idx + 1, column=1, value=value)
            sheet.cell(row=idx + 1, column=1).alignment = Alignment(horizontal="center")  # Alignment tengah

        # Menulis header summary_df_2 di bawah title
        header_row = len(title) + 1
        for col_idx, header in enumerate(summary_df_2.columns, start=1):
            sheet.cell(row=header_row, column=col_idx, value=header)
            sheet.cell(row=header_row, column=col_idx).alignment = Alignment(horizontal="center")

        # Menulis data summary_df_2 di bawah header
        for r_idx, row in enumerate(summary_df_2.itertuples(index=False), start=header_row + 1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value if not pd.isnull(value) else "")

        # Menyesuaikan lebar kolom
        for col_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            col_letter = col_cells[0].column_letter
            sheet.column_dimensions[col_letter].width = max_length + 2
        return sheet
    elif position=="center":
        # Bersihkan seluruh sheet
        sheet.delete_rows(1, sheet.max_row)

        # Menulis title dan melakukan merge
        max_col = len(summary_df_2.columns)
        for idx, value in enumerate(title):
            merge_start = 1
            merge_end = max_col if max_col > 1 else 1
            sheet.merge_cells(start_row=idx + 1, start_column=merge_start, end_row=idx + 1, end_column=merge_end)
            cell = sheet.cell(row=idx + 1, column=1, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Menulis header summary_df_2 di bawah title
        header_row = len(title) + 1
        for col_idx, header in enumerate(summary_df_2.columns, start=1):
            sheet.cell(row=header_row, column=col_idx, value=header)
            sheet.cell(row=header_row, column=col_idx).alignment = Alignment(horizontal="center")

        # Menulis data summary_df_2 di bawah header
        for r_idx, row in enumerate(summary_df_2.itertuples(index=False), start=header_row + 1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value if not pd.isnull(value) else "")

        # Menyesuaikan lebar kolom (dipersempit agar lebih rapi)
        for col_idx, col_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            max_length = max_length if max_length > 10 else 10  # Minimum width 10
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = max_length + 1  # Dipersempit dengan menambahkan 1 saja

        return sheet

def transform_name(sheet_name):
    r_prefixes = ["RIGHT", "Right", "right"]
    l_prefixes = ["LEFT", "Left", "left"]
    
    # Jika nama sheet diawali dengan "RIGHT" atau "LEFT"
    for prefix in r_prefixes + l_prefixes:
        if sheet_name.startswith(prefix):
            # Jika diawali dengan prefix, tambahkan "SD_" setelah prefix
            return sheet_name.replace(prefix, prefix + "SD_", 1)
    
    # Jika tidak ada prefix, tambahkan "SD_" di depan
    return "SD_" + sheet_name

def font_set(workbook, sheet):
    # Create the style if it doesn't exist
    # arial_style = NamedStyle(name="arial_style")
    # arial_style.font = Font(name='Arial', size=11)

    # # Only add the style if it does not already exist in the workbook
    # if arial_style.name not in workbook.named_styles:
    #     workbook.add_named_style(arial_style)

    # # Apply the style to each cell in the sheet
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         cell.style = arial_style

    return sheet




def process_csv(file_path, selected_headers, divider, firstvalue, encoding,idinput):

    startpoint=""
    endpoint=""
    # Inisialisasi list untuk menyimpan kota yang dilewati
    kotadilewati = []
    pairs = generate_pairs(selected_headers)

    try:
        trend="up"
        start_row, rows_to_read = automatevalidrow(file_path, selected_headers, encoding)
        rows_to_skip = [0] + list(range(2, start_row + 2))  # Mulai dari 2 sampai start_row + 1


        data = pd.read_csv(file_path, sep=',', encoding=encoding, skiprows=rows_to_skip,nrows=rows_to_read)
        lines=data.iloc[2]['Line']

        def loopcheck(data, startloop):
            startloop += 1
            if startloop >= len(data):
                raise IndexError("Index `start` melebihi jumlah baris pada data.")



            firstvalue =  int(data.iloc[startloop]['Km'])# meter

            meter_start = int(data.iloc[startloop]['Km']*1000 - int(data.iloc[startloop]['Km']) * 1000)

            checkvalue = int(data.iloc[startloop]['Km'] * 10000) % 10



            if(data.iloc[startloop]['Km']>data.iloc[startloop+1]['Km']):
                trend="down"
            else:
                trend="up"

            return firstvalue, meter_start, checkvalue, startloop,trend

        meter_start, checkvalue, startloop = 0, 0, -1

        if firstvalue == "":
            try:
                firstvalue, meter_start, checkvalue, startloop,trend = loopcheck(data, startloop)
                max_iterations = len(data)
                while not (1.9 <= checkvalue <= 3.1) and startloop < max_iterations:
                    firstvalue, meter_start, checkvalue, startloop,trend = loopcheck(data, startloop)
                    print("Ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
                if startloop >= max_iterations:
                    print("Tidak ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
            except IndexError:
                print("Tidak ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
        else:
            firstvalue = float(firstvalue)

        rows_to_skip = [0] + list(range(2, start_row + 2+startloop))  # Mulai dari 2 sampai start_row + 1

        data = pd.read_csv(file_path, sep=',', encoding=encoding, skiprows=rows_to_skip,nrows=rows_to_read-startloop)

        error_string = 'raise error'
        error_rows = data[data.apply(lambda row: row.astype(str).str.contains(error_string).any(), axis=1)].index.tolist()
        if error_rows:
            raise ValueError(f"Data mengandung string '{error_string}' pada baris ke: {error_rows}")

        headers = data.columns.tolist()
        print(headers)

        # Filter kolom berdasarkan header yang dipilih
        selected_columns = data[selected_headers]

        # Validasi nilai divider
        if divider <= 0:
            raise ValueError("Divider harus berupa bilangan positif.")

        # Validasi data kosong setelah filter
        if selected_columns.empty:
            raise ValueError("Data kosong setelah filtering berdasarkan header yang dipilih.")

        # Proses data sesuai kebutuhan

        # Menghapus baris pertama dari data untuk isi sheet
        data_to_save = selected_columns.iloc[1:]

        # Menyimpan hasil ke file Excel baru
        output_file = "allparameter.xlsx"
        stdev_summary = {}  # Untuk menyimpan nilai standar deviasi setiap sheet

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Menyimpan semua kolom ke satu sheet bernama "AllParameter"
            data_to_save.to_excel(writer, index=False, header=True, sheet_name="AllParameter")

            # Definisi parameter dan pasangan berdasarkan nama
            parameters = []
            # Menyimpan setiap kolom ke sheet terpisah dengan nama sesuai header
            for idx, header in enumerate(selected_headers):
               
                sheet_name_cleaned = re.sub(r'\[.*?\]', '', header)  # Menghapus teks dalam tanda kurung siku
                sheet_name_cleaned = re.sub(r'[.\-]', '', sheet_name_cleaned).strip()  # Menghapus titik dan tanda minus

                # Potong nama sheet jika lebih dari 31 karakter
                sheet_name = sheet_name_cleaned[:31] if sheet_name_cleaned else f"Sheet{idx + 1}"

                sheet_name = transform_name(sheet_name)

                parameters.append(sheet_name)

                selected_column_data = data.iloc[1:, [data.columns.get_loc(header)]]  # Data dari baris ke-2 dan seterusnya

                # Membagi data menjadi blok dengan divider baris
                total_rows = len(selected_column_data)
                blocks = [selected_column_data.iloc[i:i + divider] for i in range(0, total_rows, divider)]

                # Membuat DataFrame untuk setiap blok
                max_length = max(len(block) for block in blocks)
                split_data = pd.DataFrame()

                for i, block in enumerate(blocks):
                    # Setiap blok menjadi kolom baru
                    column_data = block.values.flatten()
                    if len(column_data) < max_length:
                        # Isi kekurangan baris dengan NaN
                        column_data = list(column_data) + [None] * (max_length - len(column_data))
                    split_data[f"Part{i + 1}"] = column_data

                # Menghitung standar deviasi untuk setiap kolom
                stdev_row = split_data.apply(lambda x: np.nanstd(x.dropna()), axis=0)

                # Menyimpan standar deviasi dalam dictionary
                stdev_summary[sheet_name] = stdev_row.values

                # Menambahkan baris standar deviasi ke DataFrame
                split_data.loc[max_length] = stdev_row

                # Menyimpan data yang telah dipecah ke sheet Excel
                split_data.to_excel(writer, index=False, header=True, sheet_name=sheet_name)

            summary_df = pd.DataFrame()

            # Menambahkan data dari stdev_summary ke summary_df
            for sheet_name, stdev_values in stdev_summary.items():
                # Pastikan stdev_values adalah array
                if isinstance(stdev_values, (np.ndarray, list)):
                    summary_df[sheet_name] = stdev_values

            # Menyimpan summary ke sheet "Summary"
            summary_df.to_excel(writer, index=False, header=True, sheet_name="Summary")


            # Inisialisasi kolom total
            total_column = []
            total_column_2 = []

            # Membersihkan setiap elemen dalam pairs
            pairs_cleaned = [
                (
                    transform_name(re.sub(r'\[.*?\]', '', param1)),
                    transform_name(re.sub(r'\[.*?\]', '', param2))   
                 )
                for param1, param2 in pairs
            ]

            pairs = [
                (re.sub(r'[.\-]', '', param1).strip(), re.sub(r'[.\-]', '', param2).strip())
                for param1, param2 in pairs_cleaned
            ]

            all_params = []
            for pair in pairs:
                for param in pair:
                    all_params.append(param)

            # Menghilangkan duplikasi
            paired_parameters = list(set(all_params))


            

            # Memastikan remaining_params benar
            remaining_params = [param for param in parameters if param not in paired_parameters]

            # Dictionary untuk indeks parameter
            param_indices = {param: idx for idx, param in enumerate(parameters)}

            
            
            # Iterasi untuk menghitung total
            for idx in range(len(summary_df)):
                total_kai = 0
                total_en = 0

                # # Menghitung total untuk pasangan parameter
                for param1_name, param2_name in pairs:
                    if param1_name in param_indices and param2_name in param_indices:
                        param1_idx = param_indices[param1_name]
                        param2_idx = param_indices[param2_name]
                        avg_pair = (summary_df.iloc[idx, param1_idx] + summary_df.iloc[idx, param2_idx]) / 2

                        total_kai += avg_pair
                        total_en += avg_pair**2

                # Menghitung total untuk sisa parameter
                for param_name in remaining_params:
                    param_idx = param_indices[param_name]
                    total_kai += summary_df.iloc[idx, param_idx]
                    total_en += summary_df.iloc[idx, param_idx]**2

                # Menyimpan hasil
                sqrtotal_en = round(total_en**0.5, 2)



                total_column_2.append(sqrtotal_en)
                total_column.append(round(total_kai, 2))
            
             # Menambahkan kolom total ke DataFrame summary_df
            summary_df['TQI Total KAI'] = total_column
            summary_df['TQI Total EN'] = total_column_2


            



            

            # Menambahkan kolom total ke DataFrame dengan aturan khusus
            summary_df['Track Number'] = [lines] * len(summary_df)


            # Menambahkan kolom "Km Section" dengan aturan khusus
            if(trend=="up"):
                summary_df['Meter Awal'] = firstvalue*1000 + meter_start + summary_df.index * divider / 4 
            elif(trend=="down"):
                summary_df['Meter Awal'] = firstvalue*1000 + meter_start - summary_df.index * divider / 4 

            # Menambahkan kolom "Track Length" dengan aturan khusus
            summary_df['Track Length'] = [divider / 4] * len(summary_df)


             # Menggabungkan kolom 'LPROF_MID' dan 'RPROF_MID' dengan menghitung rata-ratanya
            if 'Meter Awal' in summary_df.columns and 'Track Length' in summary_df.columns:
                if(trend=="up"):
                    summary_df['Meter Akhir'] = (summary_df['Meter Awal'] + summary_df['Track Length']) 
                elif(trend=="down"):
                    summary_df['Meter Akhir'] = (summary_df['Meter Awal'] - summary_df['Track Length']) 

            if 'Meter Awal' in summary_df.columns:
                summary_df['Km Awal'] = np.floor(summary_df['Meter Awal'] / 1000).astype(int) 

            if 'Meter Akhir' in summary_df.columns:
                summary_df['Km Akhir'] = np.floor(summary_df['Meter Akhir'] / 1000).astype(int) 


            if 'Meter Awal' in summary_df.columns:
                summary_df['Meter Awal'] = summary_df['Meter Awal'] -summary_df['Km Awal']*1000

            if 'Meter Akhir' in summary_df.columns:
                summary_df['Meter Akhir'] = summary_df['Meter Akhir']-summary_df['Km Akhir']*1000


            # Menentukan urutan kolom yang diinginkan
            desired_columns = (
                ['Track Number','Km Awal', 'Meter Awal','Km Akhir', 'Meter Akhir', 'Track Length'] +  # Kolom awal
                list(stdev_summary.keys()) +  # Kolom dari sheet_name (dinamis)
                ['TQI Total KAI','TQI Total EN']  # Kolom terakhir
            )

            # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df = summary_df[desired_columns]

            # Modifikasi index dengan mengalikan index dengan divider
            summary_df.index = divider / 4* (summary_df.index+1) 

            kota = []
            if(lines=='Line 1'):
                # Base data untuk Petak Jalan
                kota = [
                    {"nama": "Harjamukti", "posisi": 14524},
                    {"nama": "Ciracas", "posisi": 8864},
                    {"nama": "Kampung Rambutan", "posisi": 7262},
                    {"nama": "TMII", "posisi": 5314},
                    {"nama": "TITIK 0", "posisi": 0},
                ]
            elif(lines=='Line 2'):
                # Base data untuk Petak Jalan
                kota = [
                    {"nama": "Dukuh Atas", "posisi": 9854},
                    {"nama": "Setiabudi", "posisi": 9060},
                    {"nama": "Rasuna Said", "posisi": 7664},
                    {"nama": "Kuningan", "posisi": 6865},
                    {"nama": "Pancoran", "posisi": 4305},
                    {"nama": "Cikoko", "posisi": 2221},
                    {"nama": "Ciliwung", "posisi": 1445},
                    {"nama": "Cawang", "posisi": 187},
                    {"nama": "TITIK 0", "posisi": 0},
                ]
            elif(lines=='Line 3'):
                # Base data untuk Petak Jalan
                kota = [
                    {"nama": "Jati Mulya", "posisi": 17405},
                    {"nama": "Bekasi Barat", "posisi": 13701},
                    {"nama": "Cikunir 2", "posisi": 10437},
                    {"nama": "Cikunir 1", "posisi": 9170},
                    {"nama": "Jatibening Baru", "posisi": 6501},
                    {"nama": "Halim", "posisi": 1363},
                    {"nama": "TITIK 0", "posisi": 0},
                ]
            # Base data untuk Petak Jalan

            # Fungsi untuk membuat base_data dari daftar kota
            def buat_base_data(kota):
                base_data = []
                for i in range(len(kota) - 1):
                    base_data.append({
                        "nama": f"{kota[i]['nama']}-{kota[i+1]['nama']}",
                        "first": kota[i]["posisi"],
                        "end": kota[i+1]["posisi"]
                    })
                return base_data

            # Membuat base_data
            base_data = buat_base_data(kota)

            # Menentukan `datakota` berdasarkan `trend`
            datakota = []
            for item in base_data:
                if trend == "up":
                    nama_split = item["nama"].split("-")
                    nama_balik = f"{nama_split[1]}-{nama_split[0]}"
                    datakota.append({
                        "nama": nama_balik,
                        "first": item["end"],
                        "end": item["first"]
                    })
                else:
                    datakota.append(item)


            

            # Menghitung awal dan akhir berdasarkan Km dan Meter
            awalPetakJalan = summary_df['Km Awal'] * 1000 + summary_df['Meter Awal']
            akhirPetakJalan = summary_df['Km Akhir'] * 1000 + summary_df['Meter Akhir']
            summary_df['NilaiTengah'] = (awalPetakJalan + akhirPetakJalan) / 2

            # Loop untuk memeriksa apakah titik stasiun ada di antara awal dan akhir
            for index, row in summary_df.iterrows():
                kota_terdeteksi = False  # Menandai apakah kota ditemukan pada baris ini
                
                for kota in datakota:
                    # Memastikan logika dinamis berdasarkan trend
                    lower_bound = min(kota["first"], kota["end"])
                    upper_bound = max(kota["first"], kota["end"])
                    
                    if lower_bound <= row['NilaiTengah'] <= upper_bound:
                        
                        # Menambahkan nama kota pada kolom 'Kota' di DataFrame summary_df
                        summary_df.loc[index, 'Petak Jalan'] = kota["nama"]
                        kota_terdeteksi = True
                        kotadilewati.append(kota["nama"])
                        break  # Jika kota ditemukan, hentikan pencarian lebih lanjut

                # Jika tidak ada kota yang terdeteksi, set "Tidak terdefinisi"
                if not kota_terdeteksi:
                    summary_df.loc[index, 'Petak Jalan'] = "Tidak terdefinisi"

            
            # Menghapus kolom 'NilaiTengah' setelah selesai digunakan
            summary_df = summary_df.drop(columns=['NilaiTengah'])

            # Menyimpan DataFrame StDevSummary ke sheet
            summary_df.to_excel(writer, index=False, header=True, sheet_name="StDevSummary")




            # Membuat salinan DataFrame summary_df
            # Membuat salinan dataframe untuk diproses
            summary_df_2 = summary_df.copy()

            # Menghapus kolom 'TQI Total EN' dari summary_df_2 jika ada
            if 'TQI Total EN' in summary_df_2.columns:
                summary_df_2.drop(columns=['TQI Total EN'], inplace=True)
                
            combined_columns = []
            def process_pairs(summary_df_2, pairs):
                # Daftar awalan untuk R dan L dengan prioritas
                r_prefixes = ["RIGHT", "Right", "right", "R", "r"]
                l_prefixes = ["LEFT", "Left", "left", "L", "l"]

                combined_columns = []

                for left_col, right_col in pairs:
                    # Cek apakah kedua kolom ada dalam dataframe
                    if left_col in summary_df_2.columns and right_col in summary_df_2.columns:
                        new_column_name = None

                        # Jika pasangan diawali dengan awalan "L" dan "R", deteksi dan buat nama kolom
                        for r_prefix in r_prefixes:
                            for l_prefix in l_prefixes:
                                # Jika kolom kiri dimulai dengan "L" dan kolom kanan dengan "R", atau sebaliknya
                                if left_col.startswith(l_prefix) and right_col.startswith(r_prefix):
                                    # Hapus prefix "L" atau "R" dari kolom kiri
                                    new_column_name = left_col[len(l_prefix):]
                                    break

                        # Jika tidak menggunakan awalan "L" dan "R", buat nama kolom berdasarkan pasangan lainnya
                        if not new_column_name:
                            new_column_name = f"{left_col}_{right_col}"  # Nama kolom berdasarkan pasangan yang lebih umum

                        # Tambahkan nama kolom baru ke dalam daftar combined_columns
                        combined_columns.append(new_column_name)

                        # Gabungkan kedua kolom dengan menghitung rata-rata dan beri nama kolom baru
                        summary_df_2[new_column_name] = (summary_df_2[left_col] + summary_df_2[right_col]) / 2

                        # Hapus kolom lama setelah rata-rata dihitung
                        summary_df_2.drop(columns=[left_col, right_col], inplace=True, errors='ignore')
                    else:
                        print(f"Kolom {left_col} atau {right_col} tidak ditemukan dalam dataframe.")  # Debugging jika kolom hilang

                return summary_df_2, combined_columns

            # Proses pasangan
            summary_df_2, combined_columns = process_pairs(summary_df_2, pairs)


            # Menentukan urutan kolom yang diinginkan
            desired_columns_2 = (
                ['Track Number','Km Awal','Meter Awal','Km Akhir', 'Meter Akhir', 'Track Length'] +  # Kolom awal
                combined_columns +  # Kolom yang sudah digabungkan berdasarkan pasangan
                remaining_params +  # Kolom yang tidak memiliki pasangan
                ['TQI Total KAI']  # Kolom terakhir
            )

            # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df_2 = summary_df_2[desired_columns_2]

            

            summary_df_2['Petak Jalan'] = summary_df['Petak Jalan']

            # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df_2 = summary_df_2.round(2)

            # Menyimpan DataFrame StDevSummary_2 ke sheet
            summary_df_2.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_2 (KAI)")




            # Membuat salinan DataFrame
            summary_df_3 = pd.DataFrame()
            endvalue = summary_df_2['Km Awal'].iloc[-1]  # Nilai terakhir dari kolom 'Km Awal'

            startpoint = kotadilewati[0].split('-')[0]
            endpoint = kotadilewati[-1].split('-')[1]


            # Menambahkan title manual sebelum header dan nilai-nilai
            title = [
                "DATA TRACK QUALITY INDEX (SESUAI PERDIR NOMOR : PER.U/KI.205/XII/1/KA-2023 )",
                f"TITIK AWAL: {startpoint}",
                f"KM AWAL: {firstvalue}",
                f"TITIK AKHIR: {endpoint}",
                f"KM AKHIR: {endvalue}",
                f"TRACK NUMBER: {lines}",
                f"ID: {idinput}",
                ""  # Baris kosong
            ]

            # Menulis DataFrame summary_df_3 ke dalam sheet "StDevSummary_3"
            summary_df_3.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_3 (KAI)")

            

           



            


            # Akses workbook dan sheet setelah menulis DataFrame
            workbook = writer.book
            sheet = workbook["StDevSummary_3 (KAI)"]

            sheet=titlecreate(sheet,title,summary_df_2,'center')
            
            sheet=wrapsheet(sheet,8,15,20)
            
            sheet=center_columns(sheet)

            sheet= font_set(workbook,sheet)



            


            indexname = ['I', 'II', 'III', 'IV']

            for i in range(len(indexname)):
                # Membuat DataFrame dengan 5 kolom kosong
                columns = [
                    'BATAS ATAS TRACK QUALITY (VERY GOOD)',
                    'BATAS ATAS TRACK QUALITY (GOOD)',
                    'BATAS ATAS TRACK QUALITY (FAIR)',
                    'BATAS ATAS TRACK QUALITY (POOR)',
                    'BATAS ATAS TRACK QUALITY (VERY POOR)'
                ]

                summary_temporary = pd.DataFrame(columns=columns)

                # Menghitung nilai sementara berdasarkan tqi_kelas
                temporaryvalue = [tqi_kelas(tqi, i, divider/4) for tqi in summary_df['TQI Total KAI']]
                baris1=0
                baris2=0
                baris3=0
                baris4=0
                baris5=0
                for arrayunique in temporaryvalue:
                    baris1+=arrayunique[0]
                    baris2+=arrayunique[1]
                    baris3+=arrayunique[2]
                    baris4+=arrayunique[3]
                    baris5+=arrayunique[4]

                # Menyusun baris-baris data yang diperlukan
                rows = [
                    [np.nan] * 5,  # Baris kosong
                    [f"RESUME KELAS JALAN {indexname[i]}","","","",""],
                    [
                    
                     
                     
                    f'TOTAL PANJANG TRACK QUALITY (VERY GOOD) (M) TQI < {25+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (GOOD) (M) {25+5*i} <= TQI < {40+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (FAIR) (M) {40+5*i} <= TQI < {55+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (POOR) (M) {55+5*i} <= TQI < {70+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (VERY POOR) (M) {70+5*i} >= TQI',
                     
                     ],
                    [baris1,baris2,baris3,baris4,baris5],
                    [np.nan] * 5,
                    [f"KELAS JALAN {indexname[i]}","","","",""],
                    [
                    f'BATAS ATAS TRACK QUALITY (VERY GOOD) TQI < {25+5*i}',
                    f'BATAS ATAS TRACK QUALITY (GOOD) {25+5*i} <= TQI < {40+5*i}',
                    f'BATAS ATAS TRACK QUALITY (FAIR) {40+5*i} <= TQI < {55+5*i}',
                    f'BATAS ATAS TRACK QUALITY (POOR) {55+5*i} <= TQI < {70+5*i}',
                    f'BATAS ATAS TRACK QUALITY (VERY POOR) {70+5*i} >= TQI',
                    ],
                ]

                # Menggabungkan baris-baris menjadi satu DataFrame
                for row in rows:
                    summary_temporary = pd.concat([summary_temporary, pd.DataFrame([row], columns=columns)], ignore_index=True)

                

                # Menambahkan hasil perhitungan ke DataFrame
                summary_temporary = pd.concat([summary_temporary, pd.DataFrame(temporaryvalue, columns=columns)], ignore_index=True)

                
                # Menyimpan ke dalam file Excel
                summary_temporary.to_excel(
                    writer, 
                    index=False, 
                    header=False, 
                    startrow=2, 
                    sheet_name=f"SL KELAS JALAN {indexname[i]}"
                )

                
                # Akses workbook dan sheet setelah menulis DataFrame
                workbook = writer.book
                sheetkhusus = workbook[f"SL KELAS JALAN {indexname[i]}"]
                # Menambahkan title manual sebelum header dan nilai-nilai
                title_khusus = [
                    "KLASIFIKASI NILAI TQI  TERHADAP TRACK QUALITY & KELAS JALAN ",
                    f"SESUAI PERDIR NOMOR : PER.U/KI.205/XII/1/KA-2023 ",
                ]
                
                sheetkhusus=titlecreate(sheetkhusus,title_khusus,summary_temporary,'center')
                sheetkhusus=wrapsheet(sheetkhusus,5,15,30)
                sheetkhusus=center_columns(sheetkhusus)

                sheetkhusus= font_set(workbook,sheetkhusus)
                


                # Hapus baris ke-3
                # Hapus baris ke-3
                sheetkhusus.delete_rows(3)

                # Merge baris ke-4 pada kolom A hingga E
                sheetkhusus.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
                sheetkhusus.cell(row=4, column=1, value=f"RESUME KELAS JALAN {indexname[i]}")

                # Atur alignment tengah (horizontal dan vertikal)
                sheetkhusus.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")

                # Merge baris ke-8 pada kolom A hingga E
                sheetkhusus.merge_cells(start_row=8, start_column=1, end_row=8, end_column=5)
                sheetkhusus.cell(row=8, column=1, value=f"KELAS JALAN {indexname[i]}")

                # Atur alignment tengah untuk baris ke-8
                sheetkhusus.cell(row=8, column=1).alignment = Alignment(horizontal="center", vertical="center")


            for i in range(len(indexname)):
                # Definisi kolom DataFrame
                columns = [
                    'LINTAS', 'TRACK NUMBER', 'KM AWAL (METER)', 'KM AKHIR (METER)', 'PANJANG (METER)',

                    f'TOTAL PANJANG TRACK QUALITY (VERY GOOD) (M) TQI < {25+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (GOOD) (M) {25+5*i} <= TQI < {40+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (FAIR) (M) {40+5*i} <= TQI < {55+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (POOR) (M) {55+5*i} <= TQI < {70+5*i}',
                    f'TOTAL PANJANG TRACK QUALITY (VERY POOR) (M) {70+5*i} >= TQI',
                    'TQI'
                ]

                rows = []
                for kota in datakota:
                    # Memfilter DataFrame berdasarkan Petak Jalan
                    filtered_df = summary_df[summary_df['Petak Jalan'] == kota["nama"]]
                    
                    # Jika DataFrame kosong, lewati iterasi ini
                    if filtered_df.empty:
                        print(f"Petak Jalan untuk kota {kota['nama']} tidak ditemukan, melewati iterasi.")
                        continue

                    # Menghitung nilai sementara berdasarkan tqi_kelas
                    temporaryvalue = [tqi_kelas(tqi, i, divider / 4) for tqi in filtered_df['TQI Total KAI']]

                    # Inisialisasi variabel akumulasi
                    baris1, baris2, baris3, baris4, baris5 = 0, 0, 0, 0, 0
                    for arrayunique in temporaryvalue:
                        baris1 += arrayunique[0]
                        baris2 += arrayunique[1]
                        baris3 += arrayunique[2]
                        baris4 += arrayunique[3]
                        baris5 += arrayunique[4]

                    # Perhitungan KM Awal dan KM Akhir
                    kmawal = int((filtered_df['Km Awal'].iloc[0] * 1000) + filtered_df['Meter Awal'].iloc[0])
                    kmakhir = int((filtered_df['Km Akhir'].iloc[-1] * 1000) + filtered_df['Meter Akhir'].iloc[-1])

                    # Total panjang dan rata-rata TQI
                    sumbaris = baris1 + baris2 + baris3 + baris4 + baris5
                    tqiakhir = sum(filtered_df['TQI Total KAI'])
                    tqiakhiraverage = tqiakhir / len(filtered_df['TQI Total KAI'])

                    # Menyusun baris data
                    rows.append([
                        kota["nama"], lines, kmawal, kmakhir, 
                        sumbaris, baris1, baris2, baris3, baris4, baris5,round(tqiakhiraverage, 2)
                    ])
    
                # Membuat DataFrame sementara
                summary_temporary_2 = pd.DataFrame(rows, columns=columns)

                # Menyimpan ke Excel
                summary_temporary_2.to_excel(
                    writer,
                    index=False,
                    header=True,
                    sheet_name=f"SR KELAS JALAN {indexname[i]}"
                )

                

                # Akses workbook dan sheet untuk modifikasi lebih lanjut
                workbook = writer.book
                sheetkhusus_2 = workbook[f"SR KELAS JALAN {indexname[i]}"]

                # Menambahkan judul manual sebelum header
                title_khusus_2 = ["SUMMARY REPORT",""]
                
                sheetkhusus_2 = center_columns(sheetkhusus_2)
                sheetkhusus_2 = titlecreate(sheetkhusus_2, title_khusus_2, summary_temporary_2,'center')
                sheetkhusus_2=wrapsheet(sheetkhusus_2,3,15,20)
                sheetkhusus_2=center_columns(sheetkhusus_2)

                sheetkhusus_2= font_set(workbook,sheetkhusus_2)


                

            summary_df_2_en = summary_df_2.copy()

            # Menghapus kolom 'TQI Total EN' dari summary_df_2 jika ada
            if 'TQI Total KAI' in summary_df_2_en.columns:
                summary_df_2_en.drop(columns=['TQI Total KAI'], inplace=True)
            
            summary_df_2_en['TQI Total EN'] = summary_df['TQI Total EN']

            # Menghapus kolom 'NilaiTengah' setelah selesai digunakan
            summary_df_2_en = summary_df_2_en.drop(columns=['Petak Jalan'])

            summary_df_2_en['Petak Jalan'] = summary_df['Petak Jalan']

            # Menyimpan ke Excel
            summary_df_2_en.to_excel(
                writer,
                index=False,
                header=True,
                sheet_name=f"StDevSummary_2 (EN)"
            )

            summary_df_3_en = summary_df_2_en.copy()
            
            # Menambahkan title manual sebelum header dan nilai-nilai
            title3= [
                "DATA TRACK QUALITY INDEX (SESUAI EN 13848-6 CoSD )",
                f"TITIK AWAL: {startpoint}",
                f"KM AWAL: {firstvalue}",
                f"TITIK AKHIR: {endpoint}",
                f"KM AKHIR: {endvalue}",
                f"TRACK NUMBER: {lines}",
                f"ID: {idinput}",
                ""  # Baris kosong
            ]

            # Menulis DataFrame summary_df_3 ke dalam sheet "StDevSummary_3"
            summary_df_3_en.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_3 (EN)")

            
            # Akses workbook dan sheet setelah menulis DataFrame
            workbook = writer.book
            sheet3 = workbook["StDevSummary_3 (EN)"]

            sheet3=titlecreate(sheet3,title3,summary_df_3_en,'center')
            sheet3=wrapsheet(sheet3,8,15,20)
            sheet3=center_columns(sheet3)
            sheet3= font_set(workbook,sheet3)

            

            # Menambahkan grafik ke dalam sheet StDevSummary
            for idx, column in enumerate(summary_df.columns):
                fig, ax = plt.subplots()
                ax.plot(summary_df.index, summary_df[column], label=column)
                ax.set_title(f"Grafik {column}")
                ax.set_xlabel('Km')
                ax.set_ylabel('TQI')
                ax.legend()

                # Simpan grafik ke dalam buffer
                img_stream = BytesIO()
                fig.savefig(img_stream, format='png')
                img_stream.seek(0)

                # Masukkan gambar ke dalam worksheet
                image = Image(img_stream)
                sheet = writer.sheets["StDevSummary"]
                image.anchor = f"A{len(summary_df) + idx + 2}"  # Mengatur posisi gambar di bawah tabel
                sheet.add_image(image)


        # Contoh penggunaan
        xlsx_path = 'allparameter.xlsx'  # Ganti dengan path file Excel Anda
        

        title= [
            "DATA TRACK QUALITY INDEX",
            "DATA TRACK QUALITY INDEX (SESUAI PERDIR NOMOR : PER.U/KI.205/XII/1/KA-2023 )",
            f"TITIK AWAL: {startpoint}",
            f"KM AWAL: {firstvalue}",
            f"TITIK AKHIR: {endpoint}",
            f"KM AKHIR: {endvalue}",
            f"TRACK NUMBER: {lines}",
            f"ID: {idinput}",
            ""  # Baris kosong
        ]
        
        sheet_name = "StDevSummary_3 (KAI)"  # Ganti dengan nama sheet yang ingin diekspor
        output_pdf_path = 'TQI Summary Report KAI.pdf'  # Path untuk menyimpan PDF
        
        export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path,title,8,[], include_header=True)

        sheet_name = "StDevSummary_3 (EN)"  # Ganti dengan nama sheet yang ingin diekspor
        output_pdf_path = 'TQI Summary Report EN.pdf'  # Path untuk menyimpan PDF

        title_2= [
            "DATA TRACK QUALITY INDEX",
            "DATA TRACK QUALITY INDEX (SESUAI EN 13848-6 CoSD )",
            f"TITIK AWAL: {startpoint}",
            f"KM AWAL: {firstvalue}",
            f"TITIK AKHIR: {endpoint}",
            f"KM AKHIR: {endvalue}",
            f"TRACK NUMBER: {lines}",
            f"ID: {idinput}",
            ""  # Baris kosong
        ]
        
        export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path,title_2,8,[], include_header=True)



       
        indexname = ['I', 'II', 'III', 'IV']
        for i in range(len(indexname)):
            sheet_name = f"SR KELAS JALAN {indexname[i]}"
            output_pdf_path = f"SR KELAS JALAN {indexname[i]}.pdf"
            title = [f"SR KELAS JALAN {indexname[i]}"]
            export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path, title, 2, [], include_header=True)

            sheet_name = f"SL KELAS JALAN {indexname[i]}"
            output_pdf_path = f"SL KELAS JALAN {indexname[i]}.pdf"
            title = [f"SUMMARY REPORT"]
            export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path, title, None, [[4, 7], [9, 'endrow']], include_header=True)

        return "Sukses"

    except Exception as e:
        return f"Terjadi kesalahan: {str(e)}"

# Daftar divider
divider_values = [str(i) for i in range(20, 101, 5)]  # List: "20", "25", "30", ..., "100"
divider_values.extend([str(200), str(1000)])

# Layout GUI
layout = [
    [
        sg.Column([  
            [sg.Text("-----------PERHATIAN-----------")],

            # [sg.Text("Pastikan nama parameter hanya terdiri atas huruf dan angka, tanpa spasi atau tanda baca lainnya.")],
            # [sg.Text("Misal : VecolictyD1, PRESSUREN2, Velocity")],

            [sg.Text("Gunakan awalan LEFT dan RIGHT jika ada parameter yang berpasangan.")],
            [sg.Text("Contoh: gunakan LEFTvelocity1 dan RIGHTvelocity1.")],
            [sg.Text("Sistem akan otomatis mengenali dan memproses pasangan parameter tersebut.")],
            [sg.Text("Pilih file CV:")],
            [sg.Input(key="-FILE-"), sg.FileBrowse(file_types=(("CSV Files", "*.csv"),))],
            [sg.Button("Preview Header")],
            [sg.Text("Pilih parameter yang ingin diambil:")],
            [sg.Listbox(values=[], select_mode="multiple", size=(50, 10), key="-HEADERS-")],

            [sg.Text("Pilih Meter:")],
            [sg.Combo(divider_values, default_value="40", key="-DIVIDER-")],
            
        ], vertical_alignment='top'),  
        sg.VerticalSeparator(),  
        sg.Column([
            
            [sg.Text("ID")],
            [sg.Input(size=(10,1), key="-ID_FIRST-"), sg.Text("-"), sg.Input(size=(10,1), key="-ID_END-")],
            [sg.Text("Pasangan Parameter:")],
            [sg.Listbox(values=[], select_mode="multiple", size=(50, 10), key="-PAIRS-")],
            [sg.Button("Proses"), sg.Button("Keluar")],
            [sg.Text("", size=(50, 2), key="-OUTPUT-")]
        ], vertical_alignment='top')  
    ]
]

# Membuat jendela GUI
window = sg.Window("TQI Calculation Report", layout)




# Event loop
while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED or event == "Keluar":
        break

    if event == "Preview Header":
        file_path = values["-FILE-"]
        if not file_path:
            window["-OUTPUT-"].update("Harap pilih file CSV terlebih dahulu.")
        else:
            try:
                def detect_encoding(file_path):
                    # Membaca file dalam mode byte untuk mendeteksi encoding
                    with open(file_path, 'rb') as f:
                        raw_data = f.read()
                        result = chardet.detect(raw_data)
                    return result['encoding']
                encoding = detect_encoding(file_path)

                data = pd.read_csv(file_path , sep=',', encoding=encoding, skiprows=1)

                headers = data.columns.tolist()
                window["-HEADERS-"].update(headers)
                window["-OUTPUT-"].update("Header berhasil dimuat. Pilih parameter yang diinginkan.")
            except Exception as e:
                window["-OUTPUT-"].update(f"Terjadi kesalahan saat membaca file: {e}")

    if event == "Proses":
        file_path = values["-FILE-"]
        selected_headers = values["-HEADERS-"]
        selected_pairs = values["-PAIRS-"]
        
        # Generate pasangan parameter berdasarkan header yang dipilih
        if selected_headers:
            pairs = generate_pairs(selected_headers)
            pair_strings = [f'{pair[0]} - {pair[1]}' for pair in pairs]
            window["-PAIRS-"].update(pair_strings)
        
        divider = int(values["-DIVIDER-"]) * 4
        firstvalue = ""                                   #values["-FIRSTVALUE-"]
        

        idinput = f"{values['-ID_FIRST-']}-{values['-ID_END-']}"


        

        if not file_path or not selected_headers:
            window["-OUTPUT-"].update("Harap masukkan file dan pilih parameter yang valid.")
        else:
            pairs = []
            for pair_str in selected_pairs:
                header1, header2 = pair_str.split(' - ')
                pairs.append((header1, header2))

            result = process_csv(file_path, selected_headers, divider, firstvalue,encoding,idinput)
            window["-OUTPUT-"].update(result)

window.close()

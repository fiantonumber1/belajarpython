import openpyxl

# Membuka file formatrencana.xlsx dan formathasil.xlsx
wb_ax = openpyxl.load_workbook('formatrencana.xlsx')
wb_bx = openpyxl.load_workbook('formathasil.xlsx')

# Mengambil sheet aktif
sheet_ax = wb_ax.active
sheet_bx = wb_bx.active

# Membaca kolom B dari formatrencana.xlsx
values_ax = []
for row in sheet_ax.iter_rows(min_col=2, max_col=2, values_only=True):
    if row[0] is not None:
        values_ax.append(row[0])

# Membaca dan menggabungkan kolom B sampai I dari formathasil.xlsx
values_bx = []
for row in sheet_bx.iter_rows(min_col=2, max_col=9, values_only=True):
    concat_value = ''.join([str(cell) if cell is not None else '' for cell in row])
    if concat_value:
        values_bx.append(concat_value)

# Membuat daftar hasil gabungan jika tidak ada di values_ax
result_list = []
for value in values_bx:
    if value not in values_ax:
        result_list.append(value)

# Menampilkan hasil dan menghitung jumlah
print("Hasil gabungan yang tidak ada di formatrencana.xlsx:")
for item in result_list:
    print(item)

print(f"\nJumlah hasil gabungan yang tidak ada di formatrencana.xlsx: {len(result_list)}")

# Menyimpan hasil ke dalam file teks
with open('hasil_gabungan_tidak_ada.txt', 'w') as file:
    file.write("Hasil gabungan yang tidak ada di formatrencana.xlsx:\n")
    for item in result_list:
        file.write(f"{item}\n")
    file.write(f"\nJumlah hasil gabungan yang tidak ada di formatrencana.xlsx: {len(result_list)}\n")

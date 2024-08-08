from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# part 1
def column_exists(sheet, col_letter):
    # Periksa apakah kolom dengan huruf col_letter ada di sheet
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        if col[0].column_letter == col_letter:
            return True
    return False

def find_batas_bawah(sheet):
    for row in sheet.iter_rows(min_row=9, values_only=True):
        if any(cell == "BATAS BAWAH" for cell in row):
            return row[0].row
    return sheet.max_row + 1

def filter_and_remove_rows(sheet, col_letter, values_to_keep):
    col_idx = None
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        if col[0].column_letter == col_letter:
            col_idx = col[0].column
            break
    
    if col_idx is None:
        print(f"Kolom {col_letter} tidak ditemukan.")
        return
    
    # Membaca semua baris kecuali header
    batas_bawah_row = find_batas_bawah(sheet)
    all_rows = list(sheet.iter_rows(min_row=9, max_row=batas_bawah_row - 1, values_only=True))
    rows_to_keep = []

    for row in all_rows:
        if row[col_idx - 1] in values_to_keep:
            rows_to_keep.append(row)

    # Clear the sheet and write the filtered rows
    sheet.delete_rows(2, sheet.max_row - 1)  # Delete all rows except header
    for row in rows_to_keep:
        sheet.append([cell for cell in row])
    
    # Delete rows below batas bawah
    if batas_bawah_row <= sheet.max_row:
        sheet.delete_rows(batas_bawah_row + 1, sheet.max_row - batas_bawah_row)

    # Concatenate values from column B to I and put the result in column A
    concatenate_columns(sheet)
    # Delete columns B to I after concatenation
    delete_columns(sheet, 'B', 'I')
    delete_columns(sheet, 'C', 'J')
    delete_columns(sheet, 'D', 'Z')

def concatenate_columns(sheet):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        concatenated_value = ''.join([str(cell.value) if cell.value is not None else '' for cell in row[1:9]])  # B to I
        if len(concatenated_value) >= 4:
            row[0].value = f'{concatenated_value[0:2]}.{concatenated_value[2:3]}-{concatenated_value[3:-2]}'
        else:
            row[0].value = "Gagal, cek manual"

def delete_columns(sheet, start_col, end_col):
    start_idx = sheet[start_col + '1'].column
    end_idx = sheet[end_col + '1'].column
    for col_idx in range(end_idx, start_idx - 1, -1):
        sheet.delete_cols(col_idx)

# Muat workbook
file_path = 'carbodykci.xlsx'
workbook = load_workbook(filename=file_path)

# Ambil nama-nama sheet
sheet_names = workbook.sheetnames

# Daftar nilai yang tidak ingin dihapus
values_to_remove = [None, 'FOR REVIEW', 'WORKING']
arraysheetprosess = ["TC1 (E121)", "M1 (E122)", 'M2 (E123)', 'T1 (E124)', 'T2 (E125)', 'T3 (E126)', 'TC2 (E127)']

# Loop untuk memeriksa setiap sheet dan melakukan filter
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    
    # Cek apakah kolom W ada
    if column_exists(sheet, 'W'):
        print(f"Memproses sheet: {sheet_name}")
        if sheet_name in arraysheetprosess:
            filter_and_remove_rows(sheet, 'W', values_to_remove)

# Hapus sheet yang tidak ada di arraysheetprosess
for sheet_name in sheet_names:
    if sheet_name not in arraysheetprosess:
        std = workbook[sheet_name]
        workbook.remove(std)

# Simpan workbook yang telah diperbarui
workbook.save('carbodykci_hasil.xlsx')
print("Proses selesai. Hasil disimpan di 'carbodykci_hasil.xlsx'.")


# Lanjutkan ke part 2
file_path = 'carbodykci_hasil.xlsx'
workbooklanjutan = load_workbook(filename=file_path)

# Buat workbook baru untuk hasil
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Hasil"

# Ambil nilai A, B, C dari setiap sheet dan simpan ke sheet baru
header = ["Sheet Name", "No Dokumen", "Nama Dokumen", "Status"]
new_sheet.append(header)

unique_a_values = set()  # Untuk menyimpan nilai unik dari kolom A
unique_b_values = set()  # Untuk menyimpan nilai unik dari kolom B
rows_to_add = []

for sheet_name in workbooklanjutan.sheetnames:
    sheet = workbooklanjutan[sheet_name]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        a_val, b_val, c_val = row[0], row[1], row[2]
        if a_val and a_val != "" and a_val != "Gagal, cek manual" and a_val is not None and b_val and b_val != "" and b_val != "Gagal, cek manual" and b_val is not None:  # Hanya jika nilai A tidak kosong dan B tidak "Gagal, cek manual"
            if b_val not in unique_b_values:
                unique_b_values.add(b_val)
                if a_val not in unique_a_values:
                    unique_a_values.add(a_val)
                    rows_to_add.append([sheet_name, a_val, b_val, c_val])

# Tambahkan baris yang sudah diproses ke sheet baru
for row in rows_to_add:
    new_sheet.append(row)

# Simpan workbook baru
new_file_path = 'hasil_lanjutan.xlsx'
new_workbook.save(new_file_path)
print(f"Part 2 selesai. Hasil disimpan di '{new_file_path}'.")

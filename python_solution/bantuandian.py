import openpyxl

def update_sheet2_with_sheet1_data(sheet1_path, sheet2_path):
    # Load workbook Excel original (mintol.xlsx)
    wb_original = openpyxl.load_workbook(sheet1_path)
    sheet2_original = wb_original['sheet2']
    sheet1_original = wb_original['sheet1']

    # Load workbook Excel update (mintol_update.xlsx)
    wb_update = openpyxl.load_workbook(sheet2_path)
    sheet2_update = wb_update['sheet2']

    # Create dictionary to store information from sheet1_original
    info_sheet1 = {}
    for row in sheet1_original.iter_rows(min_row=2, values_only=True):
        key_a = row[0]  # Column A in sheet1_original
        key_c = row[2]  # Column C in sheet1_original (release date)
        if key_a not in info_sheet1:
            info_sheet1[key_a] = key_c  # Store release date keyed by value in column A

    # Process sheet2_update
    for row in sheet2_update.iter_rows(min_row=2, values_only=True):
        b_value = row[1]  # Column B in sheet2_update
        if b_value in info_sheet1:
            tanggal_rilis = info_sheet1[b_value]
            # Update column D and E in sheet2_update
            sheet2_update.cell(row=row[0] + 1, column=4, value='RILIS')  # Column D
            sheet2_update.cell(row=row[0] + 1, column=5, value=tanggal_rilis)  # Column E

    # Save the updated workbook
    wb_update.save(sheet2_path)

    print("Data has been updated in", sheet2_path)

# Example usage:
if __name__ == "__main__":
    sheet1_path = 'mintol.xlsx'
    sheet2_path = 'mintol_update.xlsx'
    update_sheet2_with_sheet1_data(sheet1_path, sheet2_path)

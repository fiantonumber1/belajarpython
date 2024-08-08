import openpyxl

# Load workbook and sheets
wb = openpyxl.load_workbook('oleh.xlsx')

# Pastikan nama sheet benar-benar sesuai dengan yang ada di file
try:
    sheet1 = wb['Sheet1']
    sheet2 = wb['Sheet2']
    sheet3 = wb['Sheet3']
except KeyError as e:
    print(f"Error: {e}")
    exit(1)

# Process Sheet1 to remove the first 3 characters in column B
processed_values = []
for row in sheet1.iter_rows(min_row=2, min_col=2, max_col=2):
    cell = row[0]
    if cell.value:
        new_value = cell.value[3:]
        processed_values.append(new_value)
    else:
        processed_values.append('')

# Check if processed values exist in Sheet2 and Sheet3 (column F)
not_in_sheet2 = []
not_in_sheet3 = []

sheet2_values = [cell for row in sheet2.iter_rows(min_row=2, min_col=6, max_col=6) for cell in row]
sheet3_values = [cell for row in sheet3.iter_rows(min_row=2, min_col=6, max_col=6) for cell in row]

for value in processed_values:
    if value not in (cell.value for cell in sheet2_values):
        not_in_sheet2.append(value)
    if value not in (cell.value for cell in sheet3_values):
        not_in_sheet3.append(value)

# Create a new workbook for the report
report_wb = openpyxl.Workbook()
report_ws = report_wb.active
report_ws.title = 'Report'

# Add headers to the report
report_ws.append(['Value', 'Not in Sheet'])

# Add data to the report
for value in not_in_sheet2:
    report_ws.append([value, 'Sheet2'])
for value in not_in_sheet3:
    report_ws.append([value, 'Sheet3'])

# Save the report
report_wb.save('report.xlsx')

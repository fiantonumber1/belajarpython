import pandas as pd
import PySimpleGUI as sg
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image
import re
from openpyxl.styles import Alignment
import string
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import pandas as pd
from fpdf import FPDF
import chardet




#### Kodingan ini berhasil untuk sistem custom untuk data mas ibnu dan sedang mengalami penyesuaian agar bisa merespon
#### data dari pak ndaru yang baru
#### Fitur preview dan choose, otomatis logic


##update V_2 -> menangani jika inputan csv


def tqi_kelas2(tqi):
    if tqi<30:
        return "Very Good"
    elif 45>=tqi>30:
        return "Good"
    elif 60>=tqi>45:
        return "Fair"
    elif 75>=tqi>60:
        return "Poor"
    else:
        return "Very Poor"
    


# Fungsi untuk membaca sheet dan mengekspor ke PDF
def export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path, firstvalue, startpoint, endpoint, endvalue):
    # Membaca data dari sheet yang dipilih, dimulai dari baris ke-7 (index 6 pada pandas)
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, skiprows=6)
    
    # Membuat objek FPDF untuk PDF
    pdf = FPDF(orientation='L', unit='mm', format='A4')  # Mengubah orientasi ke landscape
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Menambahkan halaman pertama
    pdf.add_page()

    # Menambahkan Judul di pojok kiri atas
    pdf.set_font('Arial', 'B', 8)  # Ukuran font judul lebih kecil
    pdf.cell(0, 8, txt="DATA TRACK QUALITY INDEX", ln=True, align='L')
    pdf.cell(0, 8, txt=f"TITIK AWAL: {startpoint}", ln=True, align='L')
    pdf.cell(0, 8, txt=f"KM AWAL: {firstvalue}", ln=True, align='L')
    pdf.cell(0, 8, txt=f"TITIK AKHIR: {endpoint}", ln=True, align='L')
    pdf.cell(0, 8, txt=f"KM AKHIR: {endvalue}", ln=True, align='L')
    
    # Menambahkan spasi setelah judul
    pdf.ln(8)

    # Menambahkan header tabel (nama kolom) dengan font lebih kecil
    pdf.set_font('Arial', 'B', 8)  # Ukuran font header lebih kecil
    col_widths = [pdf.get_string_width(col) + 6 for col in df.columns]  # Lebar kolom lebih kecil
    total_width = sum(col_widths)  # Lebar total tabel

    # Membuat tabel header
    for col, width in zip(df.columns, col_widths):
        pdf.cell(width, 6, col, border=1, align='C')  # Ukuran baris lebih kecil
    pdf.ln()

    # Menambahkan isi tabel (data dari dataframe) dengan font lebih kecil
    pdf.set_font('Arial', '', 6)  # Ukuran font isi tabel lebih kecil
    for row in df.itertuples(index=False):
        for col, value in zip(df.columns, row):
            pdf.cell(col_widths[df.columns.get_loc(col)], 6, str(value), border=1, align='C')  # Ukuran baris lebih kecil
        pdf.ln()

    # Menambahkan nomor halaman di pojok kanan atas pada setiap halaman
    def footer():
        pdf.set_y(-15)  # Menempatkan nomor halaman di bagian bawah
        pdf.set_font('Arial', 'I', 6)  # Ukuran font nomor halaman lebih kecil
        pdf.cell(0, 10, f'Page {pdf.page_no()}', 0, 0, 'R')  # Nomor halaman di pojok kanan atas

    # Menambahkan footer ke setiap halaman
    pdf.footer = footer

    # Menyimpan PDF ke file output
    pdf.output(output_pdf_path)


def generate_pairs(selected_headers):
    pairs = set()  # Gunakan set untuk memastikan pasangan unik
    r_prefixes = ["RIGHT", "Right", "right", "R", "r"]
    l_prefixes = ["LEFT", "Left", "left", "L", "l"]
    
    for header in selected_headers:
        # Cek jika header dimulai dengan salah satu awalan di r_prefixes
        for r_prefix in r_prefixes:
            if header.startswith(r_prefix):
                for l_prefix in l_prefixes:
                    pair = header.replace(r_prefix, l_prefix, 1)
                    if pair in selected_headers:
                        # Urutkan pasangan sebelum dimasukkan ke dalam set
                        pairs.add(tuple(sorted((header, pair))))
                        break
                break
        
        # Cek jika header dimulai dengan salah satu awalan di l_prefixes
        for l_prefix in l_prefixes:
            if header.startswith(l_prefix):
                for r_prefix in r_prefixes:
                    pair = header.replace(l_prefix, r_prefix, 1)
                    if pair in selected_headers:
                        # Urutkan pasangan sebelum dimasukkan ke dalam set
                        pairs.add(tuple(sorted((header, pair))))
                        break
                break
    
    # Ubah set ke list agar lebih mudah digunakan
    return list(pairs)

# Fungsi untuk validasi baris yang valid
def automatevalidrow(file_path, selected_headers, encoding):
    print(encoding)
    temp_data = pd.read_csv(file_path, sep=';', encoding=encoding, skiprows=1)

    # Validasi header yang dipilih
    if not set(selected_headers).issubset(temp_data.columns):
        raise ValueError("Beberapa header yang dipilih tidak ada di file CSV.")

    # Filter kolom berdasarkan header yang dipilih
    selected_columns = temp_data[selected_headers]

    # Buang baris yang memiliki nilai 'NV' di kolom yang dipilih
    selected_columns = selected_columns[~selected_columns.apply(lambda x: x.astype(str).str.contains('NV')).any(axis=1)]

    # Identifikasi baris dengan angka di semua kolom yang dipilih
    numeric_rows = selected_columns.applymap(
        lambda x: isinstance(x, (int, float)) or (isinstance(x, str) and x.replace('.', '', 1).isdigit())
    ).all(axis=1)

    start_row = numeric_rows[numeric_rows == False].index[0] if not numeric_rows.all() else None
    print(f"Start row: {start_row}")


    if start_row is None:
        return None, 0  # Jika semua baris valid, tidak perlu lanjutkan

    # Mencari end_row (baris terakhir yang False)
    end_row = numeric_rows[numeric_rows == False].index[-1]  # Baris terakhir yang False
    print(f"End row: {end_row}")

    # Menghitung jumlah baris yang akan dibaca
    rows_to_read = end_row - start_row + 1
    print(f"Rows to read: {rows_to_read}")

    return start_row, rows_to_read




def process_csv(file_path, selected_headers, divider, lines, firstvalue, startpoint, endpoint, encoding):

    pairs = generate_pairs(selected_headers)

    try:
        trend="up"
        start_row, rows_to_read = automatevalidrow(file_path, selected_headers, encoding)
        rows_to_skip = [0] + list(range(2, start_row + 2))  # Mulai dari 2 sampai start_row + 1


        data = pd.read_csv(file_path, sep=';', encoding=encoding, skiprows=rows_to_skip,nrows=rows_to_read)

        def loopcheck(data, startloop):
            startloop += 1
            if startloop >= len(data):
                raise IndexError("Index `start` melebihi jumlah baris pada data.")



            firstvalue =  int(data.iloc[startloop]['Km'])# meter

            meter_start = int(data.iloc[startloop]['Km']*1000 - int(data.iloc[startloop]['Km']) * 1000)

            checkvalue = int(data.iloc[startloop]['Km'] * 10000) % 10



            if(data.iloc[startloop]['Km']>data.iloc[startloop+1]['Km']):
                trend="down"
            else:
                trend="up"

            return firstvalue, meter_start, checkvalue, startloop,trend

        meter_start, checkvalue, startloop = 0, 0, -1

        if firstvalue == "":
            try:
                firstvalue, meter_start, checkvalue, startloop,trend = loopcheck(data, startloop)
                max_iterations = len(data)
                while not (1.9 <= checkvalue <= 3.1) and startloop < max_iterations:
                    firstvalue, meter_start, checkvalue, startloop,trend = loopcheck(data, startloop)
                    print("Ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
                if startloop >= max_iterations:
                    print("Tidak ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
            except IndexError:
                print("Tidak ditemukan baris dengan checkvalue 1.9 <= checkvalue <= 3.1")
        else:
            firstvalue = float(firstvalue)

        rows_to_skip = [0] + list(range(2, start_row + 2+startloop))  # Mulai dari 2 sampai start_row + 1

        data = pd.read_csv(file_path, sep=';', encoding=encoding, skiprows=rows_to_skip,nrows=rows_to_read-startloop)

        error_string = 'raise error'
        error_rows = data[data.apply(lambda row: row.astype(str).str.contains(error_string).any(), axis=1)].index.tolist()
        if error_rows:
            raise ValueError(f"Data mengandung string '{error_string}' pada baris ke: {error_rows}")

        headers = data.columns.tolist()
        print(headers)

        # Filter kolom berdasarkan header yang dipilih
        selected_columns = data[selected_headers]

        # Validasi nilai divider
        if divider <= 0:
            raise ValueError("Divider harus berupa bilangan positif.")

        # Validasi data kosong setelah filter
        if selected_columns.empty:
            raise ValueError("Data kosong setelah filtering berdasarkan header yang dipilih.")

        # Proses data sesuai kebutuhan

        # Menghapus baris pertama dari data untuk isi sheet
        data_to_save = selected_columns.iloc[1:]

        # Menyimpan hasil ke file Excel baru
        output_file = "allparameter.xlsx"
        stdev_summary = {}  # Untuk menyimpan nilai standar deviasi setiap sheet

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Menyimpan semua kolom ke satu sheet bernama "AllParameter"
            data_to_save.to_excel(writer, index=False, header=True, sheet_name="AllParameter")

            # Definisi parameter dan pasangan berdasarkan nama
            parameters = []
            # Menyimpan setiap kolom ke sheet terpisah dengan nama sesuai header
            for idx, header in enumerate(selected_headers):
               
                sheet_name_cleaned = re.sub(r'\[.*?\]', '', header)  # Menghapus teks dalam tanda kurung siku
                sheet_name_cleaned = re.sub(r'[.\-]', '', sheet_name_cleaned).strip()  # Menghapus titik dan tanda minus

                # Potong nama sheet jika lebih dari 31 karakter
                sheet_name = sheet_name_cleaned[:31] if sheet_name_cleaned else f"Sheet{idx + 1}"

                parameters.append(sheet_name)

                selected_column_data = data.iloc[1:, [data.columns.get_loc(header)]]  # Data dari baris ke-2 dan seterusnya

                # Membagi data menjadi blok dengan divider baris
                total_rows = len(selected_column_data)
                blocks = [selected_column_data.iloc[i:i + divider] for i in range(0, total_rows, divider)]

                # Membuat DataFrame untuk setiap blok
                max_length = max(len(block) for block in blocks)
                split_data = pd.DataFrame()

                for i, block in enumerate(blocks):
                    # Setiap blok menjadi kolom baru
                    column_data = block.values.flatten()
                    if len(column_data) < max_length:
                        # Isi kekurangan baris dengan NaN
                        column_data = list(column_data) + [None] * (max_length - len(column_data))
                    split_data[f"Part{i + 1}"] = column_data

                # Menghitung standar deviasi untuk setiap kolom
                stdev_row = split_data.apply(lambda x: np.nanstd(x.dropna()), axis=0)

                # Menyimpan standar deviasi dalam dictionary
                stdev_summary[sheet_name] = stdev_row.values

                # Menambahkan baris standar deviasi ke DataFrame
                split_data.loc[max_length] = stdev_row

                # Menyimpan data yang telah dipecah ke sheet Excel
                split_data.to_excel(writer, index=False, header=True, sheet_name=sheet_name)

            summary_df = pd.DataFrame()

            # Menambahkan data dari stdev_summary ke summary_df
            for sheet_name, stdev_values in stdev_summary.items():
                # Pastikan stdev_values adalah array
                if isinstance(stdev_values, (np.ndarray, list)):
                    summary_df[sheet_name] = stdev_values

            # Menyimpan summary ke sheet "Summary"
            summary_df.to_excel(writer, index=False, header=True, sheet_name="Summary")

            # Menambahkan sheet baru dengan penjumlahan setiap 5 data dari stdev_summary
            # for sheet_name, stdev_values in stdev_summary.items():
            #     # Tentukan ukuran blok berdasarkan nama sheet
            #     block_size = 10 if sheet_name == "Gauge" else 5
                
            #     # Membagi array stdev menjadi blok sesuai ukuran
            #     stdev_blocks = [stdev_values[i:i + block_size] for i in range(0, len(stdev_values), block_size)]
                
            #     # Menjumlahkan setiap blok
            #     summed_stdev = [sum(block) for block in stdev_blocks]
                
            #     # Membuat DataFrame untuk setiap parameter
            #     summed_df = pd.DataFrame({sheet_name: summed_stdev})
                
            #     # Menyimpan hasil penjumlahan ke sheet baru dengan nama parameter
            #     summed_df.to_excel(writer, index=False, header=True, sheet_name=f"Summed_{sheet_name}")



            


        # Contoh penggunaan
        xlsx_path = 'allparameter.xlsx'  # Ganti dengan path file Excel Anda
        sheet_name = 'StDevSummary_3'  # Ganti dengan nama sheet yang ingin diekspor
        output_pdf_path = 'TQI Summary Report VMKU.pdf'  # Path untuk menyimpan PDF

        
        return f"Data berhasil disimpan ke {output_file}"

    except Exception as e:
        return f"Terjadi kesalahan: {str(e)}"




# Daftar divider
divider_values = [str(i) for i in range(20, 101, 5)]  # List: "20", "25", "30", ..., "100"
divider_values.extend([str(200), str(1000)])

# Layout GUI
layout = [
    [
        sg.Column([  
            [sg.Text("-----------PERHATIAN-----------")],

            # [sg.Text("Pastikan nama parameter hanya terdiri atas huruf dan angka, tanpa spasi atau tanda baca lainnya.")],
            # [sg.Text("Misal : VecolictyD1, PRESSUREN2, Velocity")],

            [sg.Text("Gunakan awalan LEFT dan RIGHT jika ada parameter yang berpasangan.")],
            [sg.Text("Contoh: gunakan LEFTvelocity1 dan RIGHTvelocity1.")],
            [sg.Text("Sistem akan otomatis mengenali dan memproses pasangan parameter tersebut.")],
            [sg.Text("Pilih file CV:")],
            [sg.Input(key="-FILE-"), sg.FileBrowse(file_types=(("CSV Files", "*.csv"),))],
            [sg.Button("Preview Header")],
            [sg.Text("Pilih parameter yang ingin diambil:")],
            [sg.Listbox(values=[], select_mode="multiple", size=(50, 10), key="-HEADERS-")],

            # [sg.Text("Pilih Meter:")],
            # [sg.Combo(divider_values, default_value="40", key="-DIVIDER-")],
            
        ], vertical_alignment='top'),  
        sg.VerticalSeparator(),  
        sg.Column([
            [sg.Text("Line")],
            [sg.Input(key="-LINES-")],
            [sg.Text("Titik Awal")],
            [sg.Input(key="-STARTPOINT-")],
            [sg.Text("Titik Akhir")],
            [sg.Input(key="-ENDPOINT-")],

            # [sg.Text("Nilai Awal (KM) (kosongi untuk otomatis)")],
            # [sg.Input(key="-FIRSTVALUE-")],



            [sg.Text("Pasangan Parameter:")],
            [sg.Listbox(values=[], select_mode="multiple", size=(50, 10), key="-PAIRS-")],
            [sg.Button("Proses"), sg.Button("Keluar")],
            [sg.Text("", size=(50, 2), key="-OUTPUT-")]
        ], vertical_alignment='top')  
    ]
]

# Membuat jendela GUI
window = sg.Window("Filter Kolom Excel", layout)




# Event loop
while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED or event == "Keluar":
        break

    if event == "Preview Header":
        file_path = values["-FILE-"]
        if not file_path:
            window["-OUTPUT-"].update("Harap pilih file Excel terlebih dahulu.")
        else:
            try:
                def detect_encoding(file_path):
                    # Membaca file dalam mode byte untuk mendeteksi encoding
                    with open(file_path, 'rb') as f:
                        raw_data = f.read()
                        result = chardet.detect(raw_data)
                    return result['encoding']
                encoding = detect_encoding(file_path)

                data = pd.read_csv(file_path , sep=';', encoding=encoding, skiprows=1)

                headers = data.columns.tolist()
                window["-HEADERS-"].update(headers)
                window["-OUTPUT-"].update("Header berhasil dimuat. Pilih parameter yang diinginkan.")
            except Exception as e:
                window["-OUTPUT-"].update(f"Terjadi kesalahan saat membaca file: {e}")

    if event == "Proses":
        file_path = values["-FILE-"]
        selected_headers = values["-HEADERS-"]
        selected_pairs = values["-PAIRS-"]
        
        # Generate pasangan parameter berdasarkan header yang dipilih
        if selected_headers:
            pairs = generate_pairs(selected_headers)
            pair_strings = [f'{pair[0]} - {pair[1]}' for pair in pairs]
            window["-PAIRS-"].update(pair_strings)
        
        divider = 160

        lines = values["-LINES-"]
        firstvalue = ""                                   #values["-FIRSTVALUE-"]
        startpoint = values["-STARTPOINT-"]
        endpoint = values["-ENDPOINT-"]

        if not file_path or not selected_headers:
            window["-OUTPUT-"].update("Harap masukkan file dan pilih parameter yang valid.")
        else:
            pairs = []
            for pair_str in selected_pairs:
                header1, header2 = pair_str.split(' - ')
                pairs.append((header1, header2))

            result = process_csv(file_path, selected_headers, divider, lines, firstvalue, startpoint, endpoint,encoding)
            window["-OUTPUT-"].update(result)

window.close()

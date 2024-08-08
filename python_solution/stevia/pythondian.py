from openpyxl import load_workbook

# Fungsi untuk memodifikasi string sesuai dengan aturan yang diberikan
def modify_string(s):
    if len(s) > 4:
        # Menambahkan "." antara anggota string ke-2 dan ke-3
        modified = s[:2] + "." + s[2:3] + "-" + s[3:-2]
        return modified
    return s  # Mengembalikan string asli jika panjangnya kurang dari 5

# Membaca file Excel menggunakan openpyxl
file_path = 'mintoldian.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# Memodifikasi kolom A
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.value is not None:  # Memeriksa apakah sel tidak kosong
            cell.value = modify_string(str(cell.value))

# Menyimpan hasil modifikasi ke file Excel baru
output_path = 'mintoldian_modified_file.xlsx'
wb.save(output_path)

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Memuat workbook dan worksheet
file_path = 'PPCW.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Mengatur format tanggal awal dari sel G3 dan G4
start_date_G3 = datetime.strptime('03-Jun-24', '%d-%b-%y')
start_date_G4 = datetime.strptime('09-Jun-24', '%d-%b-%y')

# Mendefinisikan tanggal akhir
end_date = datetime.strptime('31-Dec-25', '%d-%b-%y')

# Fungsi untuk mengisi kolom berdasarkan tanggal awal
def fill_columns(start_date, start_row):
    current_date = start_date
    col_index = 8  # Kolom H dimulai dari index 8

    while current_date <= end_date:
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}{start_row}'] = current_date.strftime('%d-%b-%y')
        current_date += timedelta(days=7)
        col_index += 1

# Mengisi kolom untuk G3 dan G4
fill_columns(start_date_G3, 3)
fill_columns(start_date_G4, 4)

# Menyimpan workbook
wb.save(file_path)

import openpyxl
import os
import re  # Import regex untuk split lebih fleksibel

# Ganti dengan path file Excel yang sesuai
file_path = 'obi.xlsx'
output_file_path = 'obi_filtered.xlsx'  # File hasil yang sudah difilter
sheet_name = 'VnV - PE'

# Cek apakah file ada
if not os.path.exists(file_path):
    raise FileNotFoundError(f"File '{file_path}' tidak ditemukan. Pastikan path sudah benar.")

# Load workbook
wb = openpyxl.load_workbook(file_path)

# Cek apakah sheet tersedia
if sheet_name not in wb.sheetnames:
    raise KeyError(f"Sheet '{sheet_name}' tidak ditemukan dalam file Excel.")

# Pilih sheet yang diinginkan
sheet = wb[sheet_name]

# Cek apakah kolom N ada
if sheet.max_column < 14:
    raise ValueError("Kolom N tidak ditemukan dalam sheet, pastikan file memiliki minimal 14 kolom.")

# Dictionary untuk menyimpan nomor unik beserta barisnya
seen_numbers = {}  # Key: nomor unik, Value: baris pertama kali ditemukan
duplicate_numbers = set()  # Untuk menyimpan nomor yang ditemukan lebih dari sekali

# Loop untuk membaca kolom N
row_data = {}
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=14, max_col=14, values_only=True), start=2):
    cell_value = row[0]
    if not cell_value or not isinstance(cell_value, str):  # Lewati jika kosong atau bukan string
        continue

    # Gunakan regex untuk split dengan spasi, newline, atau tab
    split_values = [s.strip() for s in re.split(r'[\s\n\t]+', cell_value) if s]  

    unique_values = []
    
    for value in split_values:
        if value in seen_numbers:  # Jika nomor sudah ada, masukkan ke daftar duplikat
            duplicate_numbers.add(value)
        else:
            seen_numbers[value] = row_idx  # Simpan nomor unik dengan baris pertamanya
            unique_values.append(value)

    formatted_value = " ".join(unique_values).strip()  # Gabungkan kembali dengan format yang sama
    row_data[row_idx] = formatted_value  # Simpan hasil per baris

# **Menampilkan daftar nomor yang ditemukan lebih dari sekali**
if duplicate_numbers:
    print(f"Nomor duplikat ditemukan di beberapa baris: {duplicate_numbers}")

# Buat workbook baru untuk menyimpan data yang sudah difilter
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = sheet_name

# Salin semua isi sheet lama ke sheet baru
for row in sheet.iter_rows(values_only=True):
    new_sheet.append(list(row))

# Perbarui kolom N dengan data unik yang sudah difilter
for row_idx, value in row_data.items():
    new_sheet.cell(row=row_idx, column=14, value=value)  # Kolom N = Kolom ke-14

# Hapus file lama jika sudah ada agar tidak error
if os.path.exists(output_file_path):
    os.remove(output_file_path)

# Simpan file baru
new_wb.save(output_file_path)
print(f"File baru telah dibuat: {output_file_path}")

# Tutup workbook
wb.close()
new_wb.close()

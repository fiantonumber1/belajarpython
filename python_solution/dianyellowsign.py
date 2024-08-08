import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook('formathasil.xlsx')
sheet = workbook.active

# Define the yellow fill
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Iterate over the rows in the sheet
for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
    cell_w = row[22]  # Column W (0-indexed, W is the 23rd column)
    cell_x = row[23]  # Column X (0-indexed, X is the 24th column)
    cell_j = row[9]   # Column J (0-indexed, J is the 10th column)

    # Check if W and X are empty and J is not empty
    if not cell_w.value and not cell_x.value and cell_j.value:
        cell_j.fill = yellow_fill

# Save the modified workbook
workbook.save('formathasilsign.xlsx')

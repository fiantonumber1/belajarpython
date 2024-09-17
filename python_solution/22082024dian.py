import openpyxl

# Load the workbook and sheets
wb = openpyxl.load_workbook('all22082024.xlsx')
sheet_tp = wb['TP']
sheet_di = wb['DI']
sheet_sm = wb['SM']

def process_sheet(tp_sheet, di_sheet):
    # Step 1: Create clearstring list by removing " - PI" and " - PT"
    clearstrings = []
    for row in tp_sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        if row[0] and (" - PI" in row[0] or " - PT" in row[0]):
            clearstring = row[0].replace(" - PI", "").replace(" - PT", "")
            clearstrings.append(clearstring)
        elif row[0]:
            clearstrings.append(row[0])

    # Step 2: Find and append missing values from DI sheet to TP sheet
    di_data = []
    for di_row in di_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if di_row[0] and di_row[0] not in clearstrings:
            di_data.append(di_row)

    # Append missing DI rows to TP sheet
    for di_row in di_data:
        tp_sheet.append(di_row)

# Process for TP-DI and TP-SM
process_sheet(sheet_tp, sheet_di)
process_sheet(sheet_tp, sheet_sm)

# Save the workbook with the updated TP sheet
wb.save('nama_file_hasil.xlsx')

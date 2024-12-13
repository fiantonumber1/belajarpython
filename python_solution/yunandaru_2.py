import pandas as pd
import PySimpleGUI as sg
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image
import re
from openpyxl.styles import Alignment
import string
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import pandas as pd
from fpdf import FPDF

def tqi_kelas2(tqi):
    if tqi<30:
        return "Very Good"
    elif 45>=tqi>30:
        return "Good"
    elif 60>=tqi>45:
        return "Fair"
    elif 75>=tqi>60:
        return "Poor"
    else:
        return "Very Poor"
    


# Fungsi untuk membaca sheet dan mengekspor ke PDF
def export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path, firstvalue, startpoint, endpoint, endvalue):
    # Membaca data dari sheet yang dipilih, dimulai dari baris ke-7 (index 6 pada pandas)
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, skiprows=6)
    
    # Membuat objek FPDF untuk PDF
    pdf = FPDF(orientation='L', unit='mm', format='A4')  # Mengubah orientasi ke landscape
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Menambahkan halaman pertama
    pdf.add_page()

    # Menambahkan Judul di pojok kiri atas
    pdf.set_font('Arial', 'B', 8)  # Ukuran font judul lebih kecil
    pdf.cell(0, 8, txt="DATA TRACK QUALITY INDEX", ln=True, align='L')
    pdf.cell(0, 8, txt=f"TITIK AWAL: {startpoint}", ln=True, align='L')
    pdf.cell(0, 8, txt=f"KM AWAL: {firstvalue}", ln=True, align='L')
    pdf.cell(0, 8, txt=f"TITIK AKHIR: {endpoint}", ln=True, align='L')
    pdf.cell(0, 8, txt=f"KM AKHIR: {endvalue}", ln=True, align='L')
    
    # Menambahkan spasi setelah judul
    pdf.ln(8)

    # Menambahkan header tabel (nama kolom) dengan font lebih kecil
    pdf.set_font('Arial', 'B', 8)  # Ukuran font header lebih kecil
    col_widths = [pdf.get_string_width(col) + 6 for col in df.columns]  # Lebar kolom lebih kecil
    total_width = sum(col_widths)  # Lebar total tabel

    # Membuat tabel header
    for col, width in zip(df.columns, col_widths):
        pdf.cell(width, 6, col, border=1, align='C')  # Ukuran baris lebih kecil
    pdf.ln()

    # Menambahkan isi tabel (data dari dataframe) dengan font lebih kecil
    pdf.set_font('Arial', '', 6)  # Ukuran font isi tabel lebih kecil
    for row in df.itertuples(index=False):
        for col, value in zip(df.columns, row):
            pdf.cell(col_widths[df.columns.get_loc(col)], 6, str(value), border=1, align='C')  # Ukuran baris lebih kecil
        pdf.ln()

    # Menambahkan nomor halaman di pojok kanan atas pada setiap halaman
    def footer():
        pdf.set_y(-15)  # Menempatkan nomor halaman di bagian bawah
        pdf.set_font('Arial', 'I', 6)  # Ukuran font nomor halaman lebih kecil
        pdf.cell(0, 10, f'Page {pdf.page_no()}', 0, 0, 'R')  # Nomor halaman di pojok kanan atas

    # Menambahkan footer ke setiap halaman
    pdf.footer = footer

    # Menyimpan PDF ke file output
    pdf.output(output_pdf_path)

# Fungsi untuk memproses file Excel
def process_excel(file_path, columns, divider,lines,firstvalue,startpoint,endpoint):
    try:
        # Membaca sheet pertama dari file Excel
        data = pd.read_excel(file_path, sheet_name=0)

        # Mengambil kolom sesuai pilihan pengguna (dikurangi 1 untuk indeks berbasis 0)
        selected_columns = data.iloc[:, [int(col) - 1 for col in columns]]

        # Menghapus baris pertama dari data untuk isi sheet
        data_to_save = selected_columns.iloc[1:]

        # Menyimpan hasil ke file Excel baru
        output_file = "allparameter.xlsx"
        stdev_summary = {}  # Untuk menyimpan nilai standar deviasi setiap sheet

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Menyimpan semua kolom ke satu sheet bernama "AllParameter"
            data_to_save.to_excel(writer, index=False, header=True, sheet_name="AllParameter")

            # Menyimpan setiap kolom ke sheet terpisah dengan nama "Sheet1", "Sheet2", dst.
            for idx, col in enumerate(columns):
                # Ambil nilai untuk nama sheet menggunakan slicing
                sheet_name_raw = f"{data.iloc[0:1, [int(col) - 1]]}"

                # Bersihkan angka, titik, tanda minus, dan trim spasi
                sheet_name_cleaned = re.sub(r'[\d.\-]', '', sheet_name_raw).strip()

                 # Pastikan sheet_name tidak kosong, gunakan nama default jika kosong
                sheet_name = sheet_name_cleaned if sheet_name_cleaned else f"Sheet{idx + 1}"


                selected_column_data = data.iloc[1:, [int(col) - 1]]  # Data dari baris ke-2 dan seterusnya
                

                # Membagi data menjadi blok divider baris
                total_rows = len(selected_column_data)
                blocks = [selected_column_data.iloc[i:i + divider] for i in range(0, total_rows, divider)]

                # Membuat DataFrame untuk setiap blok
                max_length = max(len(block) for block in blocks)
                split_data = pd.DataFrame()

                for i, block in enumerate(blocks):
                    # Setiap blok menjadi kolom baru
                    column_data = block.values.flatten()
                    if len(column_data) < max_length:
                        # Isi kekurangan baris dengan NaN
                        column_data = list(column_data) + [None] * (max_length - len(column_data))
                    split_data[f"Part{i + 1}"] = column_data

                # Menghitung standar deviasi untuk setiap kolom
                stdev_row = split_data.apply(lambda x: np.nanstd(x.dropna()), axis=0)

                # Menyimpan standar deviasi dalam dictionary
                stdev_summary[sheet_name] = stdev_row.values

                # Menambahkan baris standar deviasi ke DataFrame
                split_data.loc[max_length] = stdev_row

                # Menyimpan data yang telah dipecah ke sheet Excel
                split_data.to_excel(writer, index=False, header=True, sheet_name=sheet_name)


            summary_df = pd.DataFrame()

            # Menambahkan data dari stdev_summary ke summary_df
            for sheet_name, stdev_values in stdev_summary.items():
                # Menambahkan data ke DataFrame (asumsikan stdev_values adalah dictionary atau Series)
                summary_df[sheet_name] = stdev_values

            # Menambahkan kolom total sesuai aturan yang diberikan
            total_column = []
            total_column_2=[]

            # Iterasi untuk menghitung total sesuai dengan aturan khusus
            for idx in range(len(summary_df)):
                total_kai = 0
                total_en = 0

                # Menghitung total berdasarkan aturan yang diberikan
                if len(summary_df.columns) >= 2:
                    total_kai += (summary_df.iloc[idx, 0] + summary_df.iloc[idx, 1]) / 2  # Parameter 1 dan 2
                    total_en += ((summary_df.iloc[idx, 0] + summary_df.iloc[idx, 1]) / 2)**2       # Parameter 1 dan 2
                if len(summary_df.columns) >= 4:
                    total_kai += (summary_df.iloc[idx, 2] + summary_df.iloc[idx, 3]) / 2  # Parameter 3 dan 4
                    total_en += ((summary_df.iloc[idx, 2] + summary_df.iloc[idx, 3]) / 2)**2  # Parameter 3 dan 4

                # Untuk sisanya, dijumlahkan
                if len(summary_df.columns) > 4:
                    total_kai += summary_df.iloc[idx, 4:].sum()

                # Untuk sisanya, dijumlahkan
                if len(summary_df.columns) > 4:
                    total_en += (summary_df.iloc[idx, 4])**2
                    total_en += (summary_df.iloc[idx, 5])**2
                    total_en += (summary_df.iloc[idx, 6])**2

                sqrtotal_en=   round(total_en**(0.5), 2)

                total_column_2.append(sqrtotal_en)

                total_column.append(round(total_kai, 2))

            # Menambahkan kolom total ke DataFrame summary_df
            summary_df['TQI Total KAI'] = total_column



            # Menghitung kelas berdasarkan TQI Total KAI
            total_column_6 = [tqi_kelas2(tqi) for tqi in total_column]
            summary_df['Class'] = total_column_6



            # Menambahkan kolom total ke DataFrame summary_df
            summary_df['TQI Total EN'] = total_column_2

            # Menambahkan kolom total ke DataFrame dengan aturan khusus
            summary_df['Lines'] = [lines] * len(summary_df)

            # Menambahkan kolom "Km Section" dengan aturan khusus
            summary_df['Meter Awal'] = summary_df.index * divider / 4 + firstvalue*1000

            # Menambahkan kolom "Track Length" dengan aturan khusus
            summary_df['Track Length'] = [divider / 4] * len(summary_df)


             # Menggabungkan kolom 'LPROF_MID' dan 'RPROF_MID' dengan menghitung rata-ratanya
            if 'Meter Awal' in summary_df.columns and 'Track Length' in summary_df.columns:
                summary_df['Meter Akhir'] = (summary_df['Meter Awal'] + summary_df['Track Length']) 

            if 'Meter Awal' in summary_df.columns:
                summary_df['Km Awal'] = np.floor(summary_df['Meter Awal'] / 1000).astype(int) 

            if 'Meter Akhir' in summary_df.columns:
                summary_df['Km Akhir'] = np.floor(summary_df['Meter Akhir'] / 1000).astype(int) 


            if 'Meter Awal' in summary_df.columns:
                summary_df['Meter Awal'] = summary_df['Meter Awal'] -summary_df['Km Awal']*1000

            if 'Meter Akhir' in summary_df.columns:
                summary_df['Meter Akhir'] = summary_df['Meter Akhir']-summary_df['Km Akhir']*1000


                




            # Menentukan urutan kolom yang diinginkan
            desired_columns = (
                ['Lines','Km Awal', 'Meter Awal','Km Akhir', 'Meter Akhir', 'Track Length'] +  # Kolom awal
                list(stdev_summary.keys()) +  # Kolom dari sheet_name (dinamis)
                ['TQI Total KAI','Class', 'TQI Total EN']  # Kolom terakhir
            )

            # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df = summary_df[desired_columns]










            


            # Modifikasi index dengan mengalikan index dengan divider
            summary_df.index = divider / 4* (summary_df.index+1) 

            # Menyimpan DataFrame StDevSummary ke sheet
            summary_df.to_excel(writer, index=False, header=True, sheet_name="StDevSummary")

            # Membuat salinan DataFrame summary_df
            summary_df_2 = summary_df.copy()

            # Menghapus kolom 'TQI Total EN' dari summary_df_2 jika ada
            if 'TQI Total EN' in summary_df_2.columns:
                summary_df_2.drop(columns=['TQI Total EN'], inplace=True)

            # Menggabungkan kolom 'LPROF_MID' dan 'RPROF_MID' dengan menghitung rata-ratanya
            if 'LPROF_MID' in summary_df_2.columns and 'RPROF_MID' in summary_df_2.columns:
                summary_df_2['TQI_PROF'] = (summary_df_2['LPROF_MID'] + summary_df_2['RPROF_MID']) / 2
                

            # Memastikan kolom LPROF_MID dan RPROF_MID dihapus jika tidak lagi diperlukan
            summary_df_2.drop(columns=['LPROF_MID', 'RPROF_MID'], inplace=True, errors='ignore')

            # Menggabungkan kolom 'LPROF_MID' dan 'RPROF_MID' dengan menghitung rata-ratanya
            if 'LALIGN_MID' in summary_df_2.columns and 'RALIGN_MID' in summary_df_2.columns:
                summary_df_2['TQI_ALIGN'] = (summary_df_2['LALIGN_MID'] + summary_df_2['RALIGN_MID']) / 2


            summary_df_2['TQI_GAGE'] = summary_df_2['GAGE']
            summary_df_2['TQI_XLEVEL'] = summary_df_2['XLEVEL']
            summary_df_2['TQI_TWIST'] = summary_df_2['TWIST_LONG']

            # Memastikan kolom LPROF_MID dan RPROF_MID dihapus jika tidak lagi diperlukan
            summary_df_2.drop(columns=['GAGE', 'XLEVEL','TWIST_LONG'], inplace=True, errors='ignore')


           
                

            # Memastikan kolom LPROF_MID dan RPROF_MID dihapus jika tidak lagi diperlukan
            summary_df_2.drop(columns=['LALIGN_MID', 'RALIGN_MID'], inplace=True, errors='ignore')


            # Menentukan urutan kolom yang diinginkan
            desired_columns_2 = (
                ['Lines','Km Awal','Meter Awal','Km Akhir', 'Meter Akhir', 'Track Length'] +  # Kolom awal
                ['TQI_PROF','TQI_ALIGN','TQI_GAGE','TQI_XLEVEL','TQI_TWIST'] +  # Kolom dari sheet_name (dinamis)
                ['TQI Total KAI','Class']  # Kolom terakhir
            )

            # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df_2 = summary_df_2[desired_columns_2]

             # Memastikan kolom diurutkan sesuai dengan desired_columns
            summary_df_2 = summary_df_2.round(2)

            # Menyimpan DataFrame StDevSummary_2 ke sheet
            summary_df_2.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_2")




            # Membuat salinan DataFrame (Jadinya ditengah)
            # summary_df_3 = pd.DataFrame()
            # endvalue = summary_df_2['Km Awal'].iloc[-1]  # Nilai terakhir dari kolom 'Km Awal'

            # # Menambahkan title manual sebelum header dan nilai-nilai
            # title = [
            #     "DATA TRACK QUALITY INDEX",
            #     f"TITIK AWAL: {startpoint}",
            #     f"KM AWAL: {firstvalue}",
            #     f"TITIK AKHIR: {endpoint}",
            #     f"KM AKHIR: {endvalue}",
            #     ""  # Baris kosong
            # ]

            # # Menulis DataFrame summary_df_3 ke dalam sheet "StDevSummary_3"
            # summary_df_3.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_3")

            # # Akses workbook dan sheet setelah menulis DataFrame
            # workbook = writer.book
            # sheet = workbook["StDevSummary_3"]

            # # Hitung jumlah kolom untuk menentukan kolom tengah
            # num_columns = len(summary_df_2.columns)
            # center_col = (num_columns // 2) + 1  # Kolom tengah (1-based index)

            # # Menulis title di tengah
            # for idx, value in enumerate(title):
            #     sheet.cell(row=idx + 1, column=center_col, value=value)
            #     sheet.cell(row=idx + 1, column=center_col).alignment = Alignment(horizontal="center")  # Alignment tengah

            # # Menulis header summary_df_2 di bawah title
            # header_row = len(title) + 1
            # for col_idx, header in enumerate(summary_df_2.columns, start=1):
            #     sheet.cell(row=header_row, column=col_idx, value=header)
            #     sheet.cell(row=header_row, column=col_idx).alignment = Alignment(horizontal="center")

            # # Menulis data summary_df_2 di bawah header
            # for r_idx, row in enumerate(summary_df_2.itertuples(index=False), start=header_row + 1):
            #     for c_idx, value in enumerate(row, start=1):
            #         sheet.cell(row=r_idx, column=c_idx, value=value if not pd.isnull(value) else "")

            # # Menyesuaikan lebar kolom
            # for col_cells in sheet.columns:
            #     max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            #     col_letter = col_cells[0].column_letter
            #     sheet.column_dimensions[col_letter].width = max_length + 2




            # Membuat salinan DataFrame
            summary_df_3 = pd.DataFrame()
            endvalue = summary_df_2['Km Awal'].iloc[-1]  # Nilai terakhir dari kolom 'Km Awal'

            # Menambahkan title manual sebelum header dan nilai-nilai
            title = [
                "DATA TRACK QUALITY INDEX",
                f"TITIK AWAL: {startpoint}",
                f"KM AWAL: {firstvalue}",
                f"TITIK AKHIR: {endpoint}",
                f"KM AKHIR: {endvalue}",
                ""  # Baris kosong
            ]

            # Menulis DataFrame summary_df_3 ke dalam sheet "StDevSummary_3"
            summary_df_3.to_excel(writer, index=False, header=True, sheet_name="StDevSummary_3")

            # Akses workbook dan sheet setelah menulis DataFrame
            workbook = writer.book
            sheet = workbook["StDevSummary_3"]

            # Menulis title
            for idx, value in enumerate(title):
                sheet.cell(row=idx + 1, column=1, value=value)
                sheet.cell(row=idx + 1, column=1).alignment = Alignment(horizontal="center")  # Alignment tengah

            # Menulis header summary_df_2 di bawah title
            header_row = len(title) + 1
            for col_idx, header in enumerate(summary_df_2.columns, start=1):
                sheet.cell(row=header_row, column=col_idx, value=header)
                sheet.cell(row=header_row, column=col_idx).alignment = Alignment(horizontal="center")

            # Menulis data summary_df_2 di bawah header
            for r_idx, row in enumerate(summary_df_2.itertuples(index=False), start=header_row + 1):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value if not pd.isnull(value) else "")

            # Menyesuaikan lebar kolom
            for col_cells in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
                col_letter = col_cells[0].column_letter
                sheet.column_dimensions[col_letter].width = max_length + 2
            
            
            
            
            


            


            
            


            # Menambahkan grafik ke dalam sheet StDevSummary
            for idx, column in enumerate(summary_df.columns):
                fig, ax = plt.subplots()
                ax.plot(summary_df.index, summary_df[column], label=column)
                ax.set_title(f"Grafik {column}")
                ax.set_xlabel('Km')
                ax.set_ylabel('TQI')
                ax.legend()

                # Simpan grafik ke dalam buffer
                img_stream = BytesIO()
                fig.savefig(img_stream, format='png')
                img_stream.seek(0)

                # Masukkan gambar ke dalam worksheet
                image = Image(img_stream)
                sheet = writer.sheets["StDevSummary"]
                image.anchor = f"A{len(summary_df) + idx + 2}"  # Mengatur posisi gambar di bawah tabel
                sheet.add_image(image)


        # Contoh penggunaan
        xlsx_path = 'allparameter.xlsx'  # Ganti dengan path file Excel Anda
        sheet_name = 'StDevSummary_3'  # Ganti dengan nama sheet yang ingin diekspor
        output_pdf_path = 'TQI Summary Report VMKU.pdf'  # Path untuk menyimpan PDF

        export_xlsx_to_pdf(xlsx_path, sheet_name, output_pdf_path,firstvalue,startpoint,endpoint,endvalue)

        return f"Data berhasil disimpan ke {output_file}"

    except Exception as e:
        return f"Terjadi kesalahan: {str(e)}"



# Membuat daftar divider
divider_values = [str(i) for i in range(20, 101, 5)]  # List: "20", "25", "30", ..., "100"

# Menambahkan nilai tambahan 200 dan 1000
divider_values.extend([str(200), str(1000)])

# Layout GUI
layout = [
    [sg.Text("Pilih file Excel:")],
    [sg.Input(key="-FILE-"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
    [sg.Text("Masukkan kolom / paremeter yang ingin diambil (pisahkan dengan koma, misal: 2,4,5,6):")],
    [sg.Input(key="-COLUMNS-")],
    [sg.Text("Pilih Meter:")],
    [sg.Combo(divider_values, default_value="40", key="-DIVIDER-")],
    [sg.Text("Line")],
    [sg.Input(key="-LINES-")],
    [sg.Text("Titik Awal")],
    [sg.Input(key="-STARTPOINT-")],

    [sg.Text("Titik Akhir")],
    [sg.Input(key="-ENDPOINT-")],

    [sg.Text("Nilai Awal (KM)")],
    [sg.Input(key="-FIRSTVALUE-")],
    [sg.Button("Proses"), sg.Button("Keluar")],
    [sg.Text("", size=(50, 2), key="-OUTPUT-")]
]

# Membuat jendela GUI
window = sg.Window("Filter Kolom Excel", layout)

# Event loop
while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED or event == "Keluar":
        break

    if event == "Proses":
        file_path = values["-FILE-"]
        columns = values["-COLUMNS-"].split(",")
        divider = int(values["-DIVIDER-"]) * 4
        lines = values["-LINES-"]
        firstvalue= float(values["-FIRSTVALUE-"])

        startpoint = values["-STARTPOINT-"]
        endpoint = values["-ENDPOINT-"]

        if not file_path or not columns:
            window["-OUTPUT-"].update("Harap masukkan file dan kolom yang valid.")
        else:
            result = process_excel(file_path, columns, divider,lines,firstvalue,startpoint,endpoint)
            window["-OUTPUT-"].update(result)

window.close()

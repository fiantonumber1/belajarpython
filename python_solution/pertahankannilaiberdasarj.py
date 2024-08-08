from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def column_exists(sheet, col_letter):
    # Periksa apakah kolom dengan huruf col_letter ada di sheet
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        if col[0].column_letter == col_letter:
            return True
    return False

def filter_and_remove_rows(sheet, col_letter, values_to_keep):
    col_idx = None
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        if col[0].column_letter == col_letter:
            col_idx = col[0].column
            break
    
    if col_idx is None:
        print(f"Kolom {col_letter} tidak ditemukan.")
        return
    
    # Membaca semua baris kecuali header
    all_rows = list(sheet.iter_rows(min_row=2, values_only=False))
    rows_to_keep = []

    for row in all_rows:
        if row[col_idx - 1].value in values_to_keep:
            rows_to_keep.append(row)

    # Clear the sheet and write the filtered rows
    sheet.delete_rows(2, sheet.max_row - 1)  # Delete all rows except header
    for row in rows_to_keep:
        sheet.append([cell.value for cell in row])


# Muat workbook
file_path = 'carbodykci.xlsx'
workbook = load_workbook(filename=file_path)

# Ambil nama-nama sheet
sheet_names = workbook.sheetnames

# Daftar nilai yang tidak ingin dihapus
values_to_remove = ['', 'FOR REVIEW', 'WORKING']

# Loop untuk memeriksa setiap sheet dan melakukan filter
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    
    # Cek apakah kolom W ada
    if column_exists(sheet, 'W'):
        print(f"Memproses sheet: {sheet_name}")
        filter_and_remove_rows(sheet, 'W', values_to_remove)

# Simpan workbook yang telah diperbarui
workbook.save('carbodykci_hasil.xlsx')
print("Proses selesai. Hasil disimpan di 'carbodykci_hasil.xlsx'.")

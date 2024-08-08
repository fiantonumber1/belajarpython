from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def column_exists(sheet, col_letter):
    # Periksa apakah kolom dengan huruf col_letter ada di sheet
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        if col[0].column_letter == col_letter:
            return True
    return False

def filter_and_remove_rows(sheet, col_letter, values_to_keep):
    col_idx = None
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        if col[0].column_letter == col_letter:
            col_idx = col[0].column
            break
    
    if col_idx is None:
        print(f"Kolom {col_letter} tidak ditemukan.")
        return
    
    # Membaca semua baris kecuali header
    all_rows = list(sheet.iter_rows(min_row=2, values_only=False))
    rows_to_keep = []

    for row in all_rows:
        if row[col_idx - 1].value in values_to_keep:
            rows_to_keep.append(row)

    # Clear the sheet and write the filtered rows
    sheet.delete_rows(2, sheet.max_row - 1)  # Delete all rows except header
    for row in rows_to_keep:
        sheet.append([cell.value for cell in row])
    
    # Concatenate values from column B to I and put the result in column A
    concatenate_columns(sheet)
    # Delete columns B to I after concatenation
    delete_columns(sheet, 'B', 'I')
    delete_columns(sheet, 'C', 'J')
    delete_columns(sheet, 'D', 'Z')

def concatenate_columns(sheet):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        concatenated_value = ''.join([str(cell.value) if cell.value is not None else '' for cell in row[1:9]])  # B to I
        if len(concatenated_value) >= 4:
            row[0].value = f'{concatenated_value[0:2]}.{concatenated_value[2:3]}-{concatenated_value[3:-2]}'
        else:
            row[0].value = "Gagal, cek manual"

def delete_columns(sheet, start_col, end_col):
    start_idx = sheet[start_col + '1'].column
    end_idx = sheet[end_col + '1'].column
    for col_idx in range(end_idx, start_idx - 1, -1):
        sheet.delete_cols(col_idx)

# Muat workbook
file_path = 'carbodykci.xlsx'
workbook = load_workbook(filename=file_path)

# Ambil nama-nama sheet
sheet_names = workbook.sheetnames

# Daftar nilai yang tidak ingin dihapus
values_to_remove = ['', 'FOR REVIEW', 'WORKING']
arraysheetprosess = ["TC1 (E121)", "M1 (E122)",'M2 (E123)','T1 (E124)','T2 (E125)','T3 (E126)','TC2 (E127)']

# Loop untuk memeriksa setiap sheet dan melakukan filter
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    
    # Cek apakah kolom W ada
    if column_exists(sheet, 'W'):
        print(f"Memproses sheet: {sheet_name}")
        if sheet_name in arraysheetprosess:
            filter_and_remove_rows(sheet, 'W', values_to_remove)

# Simpan workbook yang telah diperbarui
workbook.save('carbodykci_hasil.xlsx')
print("Proses selesai. Hasil disimpan di 'carbodykci_hasil.xlsx'.")
